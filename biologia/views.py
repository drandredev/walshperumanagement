from django.shortcuts import render, redirect
from biologia.models import *
from biologia.forms import *
from django.http import HttpResponseRedirect
import sweetify
from django.contrib import messages


# Create your views here.

#Vistas a usarse()

#-----------------------------------------------------------------------------------------------------------------------

#Avez Views

def form_BD_Avez_Chinalco(request):
    datos_form=BD_Integrada_Aves_Chinalco_Seca_2018.objects.all()

    if request.method == 'POST':
        form = FormularioForm(request.POST,request.FILES)
        print(form)
        if form.is_valid():
            form.save()
            print(form)
            messages.success(request, "Formulario enviado correctamente." )
            return HttpResponseRedirect('/biologia/form_base_datos_avez/')
        messages.error(request, "Ocurrió un error, vuelve a intentarlo.")
    else:
        form = FormularioForm()

    return render(request,'biologia/formulario_Aves_chinalco_BD.html',
                  {'datos_form':datos_form,
                   'form': form})

def form_fotos_integrado_Avez_Chinalco(request):

    form_fotos_integrado=Fotos_Integrado_Aves_Chinalco_Seca_2018.objects.all()

    if request.method == 'POST':
        form = FormularioFotosIntegrado(request.POST,request.FILES)
        print(form)
        if form.is_valid():
            form.save()
            messages.success(request, "Formulario enviado correctamente." )
            return HttpResponseRedirect('/biologia/form_fotos_integrado_avez/')
        messages.error(request, "Ocurrió un error, vuelve a intentarlo.")
    else:
        form = FormularioFotosIntegrado()

    return render(request,'biologia/formulario_fotos_integrado_aves_chinalco.html',
                  {'form_fotos_integrado':form_fotos_integrado,
                   'form': form})

def form_coordenadas_aves(request):

    if request.method == 'POST':
        form = FormularioCoordenadasAves(request.POST,request.FILES)
        print(form)
        if form.is_valid():
            form.save()
            messages.success(request, "Formulario enviado correctamente." )
            return HttpResponseRedirect('/biologia/form_coordenadas_aves/')
        messages.error(request, "Ocurrió un error, vuelve a intentarlo.")
    else:
        form = FormularioCoordenadasAves()

    return render(request,'biologia/formulario_coordenadas_integradas_aves_chinalco.html',
                  {'form':form})

#-----------------------------------------------------------------------------------------------------------------------

#Reptiles_Amfibios Views

def form_anfibios_reptiles_base_datos(request):

    if request.method == 'POST':
        form = FormularioBdAnfibiosReptiles(request.POST,request.FILES)
        print(form)
        if form.is_valid():
            form.save()
            messages.success(request, "Formulario enviado correctamente." )
            return redirect('/biologia/form_base_anfibios_reptiles/')
        messages.error(request, "Ocurrió un error, vuelve a intentarlo.")
    else:
        form = FormularioBdAnfibiosReptiles()

    return render(request,'biologia/formulario_bd_anfibios_reptiles.html',
             {'form':form})

def form_coordenadas_anfibios_repites_bd(request):

    form_coordenadas_anf_rep_db = Coordenadas_BD.objects.all()

    if request.method == 'POST':
        form = FormularioCoordenadas_BD(request.POST,request.FILES)
        print(form)
        if form.is_valid():
            form.save()
            messages.success(request, "Formulario enviado correctamente." )
            return redirect('../../biologia/form_coord_anf_rep/')
        messages.error(request, "Ocurrió un error, vuelve a intentarlo.")
    else:
        form = FormularioCoordenadas_BD()

    return render(request,'biologia/formulario_bd_coordenadas.html',
                  {'form_coordenadas_anf_rep_db':form_coordenadas_anf_rep_db,
                   'form': form})

def form_db_hoja1(request):

    if request.method == 'POST':
        form = FormularioHoja1(request.POST,request.FILES)
        print(form)
        if form.is_valid():
            form.save()
            messages.success(request, "Formulario enviado correctamente." )
            return redirect('../../biologia/form_hoja_1/')
        messages.error(request, "Ocurrió un error, vuelve a intentarlo.")
    else:
        form = FormularioHoja1()

    return render(request,'biologia/formulario_db_hoja1.html',
                  {'form':form})

def form_fotos_db(request):

    if request.method == 'POST':
        form = Formulario_BD_fotos_BD(request.POST,request.FILES)
        print(form)
        if form.is_valid():
            form.save()
            messages.success(request, "Formulario enviado correctamente." )
            return redirect('../../biologia/form_fotos_bd/')
        messages.error(request, "Ocurrió un error, vuelve a intentarlo.")
    else:
        form = Formulario_BD_fotos_BD()

    return render(request, 'biologia/formulario_bd_fotos_Anfibios_reptiles.html',
                  {'form':form})


#-----------------------------------------------------------------------------------------------------------------------

#Mamiferos Views

def form_bitacora_mamiferos(request):

    form_bitacora_mamiferos = bitacora_BD_Mamiferos.objects.all()

    if request.method == 'POST':
        form = FormularioBitacora_Mamiferos(request.POST,request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Formulario enviado correctamente." )
            return redirect('../../biologia/form_bitacora_mamiferos/')
    else:
        form = FormularioBitacora_Mamiferos()

    return render(request,'biologia/formulario_bd_mamifero_bitacora.html',
                  {'form_bitacora_mamiferos':form_bitacora_mamiferos,
                   'form': form})

def form_esfuerzo_muestreo_mamiferos(request):

    if request.method == 'POST':
        form = FormularioEsfuerzoMamifero(request.POST,request.FILES)
        print(form)
        if form.is_valid():
            form.save()
            messages.success(request, "Formulario enviado correctamente." )
            return HttpResponseRedirect('/biologia/form_esfuerzo_muestreo_mamiferos/')
    else:
        form = FormularioEsfuerzoMamifero()

    return render(request,'biologia/formulario_esfuerzo_muestreo_mamiferos_chinalco.html',
                  {'form':form})

def form_coordenadas_reporte_mamiferos(request):

    if request.method == 'POST':
        form = FormularioCoordenadasMamiferos(request.POST,request.FILES)
        print(form)
        if form.is_valid():
            form.save()
            messages.success(request, "Formulario enviado correctamente." )
            return HttpResponseRedirect('/biologia/form_coordenadas_reporte_mamiferos/')
    else:
        form = FormularioCoordenadasMamiferos()

    return render(request,'biologia/formulario_coordenadas_reporte_mamiferos.html',
                  {'FormularioCoordenadasMamiferos':FormularioCoordenadasMamiferos,
                   'form': form})

def form_BD_mamiferos(request):

    if request.method == 'POST':
        form = FormularioBdMamiferos(request.POST,request.FILES)
        print(form)
        if form.is_valid():
            form.save()
            messages.success(request, "Formulario enviado correctamente." )
            return HttpResponseRedirect('../../biologia/form_BD_mamiferos/')
    else:
        form = FormularioBdMamiferos()

    return render(request,'biologia/formulario_BD_Mamiferos_Chinalco.html',
                  {'form':form})

#-----------------------------------------------------------------------------------------------------------------------

#Vegetacion Views


def form_hoja3_vegetacion(request):

    if request.method == 'POST':
        form = FormularioHoja3_Vegetacion(request.POST,request.FILES)
        print(form)
        if form.is_valid():
            form.save()
            messages.success(request, "Formulario enviado correctamente." )
            return HttpResponseRedirect('../../biologia/form_hoja3_vegetacion/')
    else:
        form = FormularioHoja3_Vegetacion()

    return render(request,'biologia/formulario_hoja3_vegetacion.html',
                  {'form':form})

def form_Vegetacion_Compensacion_Ambiental(request):

    if request.method == 'POST':
        form = Formulario_Vegetacion_CompensacionAmbiental(request.POST,request.FILES)
        print(form)
        if form.is_valid():
            form.save()
            messages.success(request, "Formulario enviado correctamente." )
            return HttpResponseRedirect('../../biologia/form_Vegetacion_Compensacion_Ambiental/')
    else:
        form = Formulario_Vegetacion_CompensacionAmbiental()

    return render(request,'biologia/formulario_vegetacion_compensacion_ambiental.html',
                  {'form':form})

def form_Vegetacion_Base_datos(request):

    if request.method == 'POST':
        form = Formulario_Vegetacion_BD(request.POST,request.FILES)
        print(form)
        if form.is_valid():
            form.save()
            messages.success(request, "Formulario enviado correctamente." )
            return HttpResponseRedirect('../../biologia/form_Vegetacion_Base_datos/')
    else:
        form = Formulario_Vegetacion_BD()

    return render(request, 'biologia/formulario_Vegetacion_Base_de_datos.html',
                  {'form':form})

def form_Vegetacion_esfuerzo_muestreo(request):

    if request.method == 'POST':
        form = Formulario_Vegetacion_EsfuerzoMuestreo(request.POST,request.FILES)
        print(form)
        if form.is_valid():
            form.save()
            messages.success(request, "Formulario enviado correctamente." )
            return HttpResponseRedirect('../../biologia/form_Vegetacion_esfuerzo_muestreo/')
    else:
        form = Formulario_Vegetacion_EsfuerzoMuestreo()

    return render(request, 'biologia/formulario_vegetacion_esfuerzo_muestreo.html',
                  {'form':form})


#-----------------------------------------------------------------------------------------------------------------------


#======================== LIst_Form views ===================

#===== AVES ====
def list_forms_Aves(request):
    lista_aves_chinalco=BD_Integrada_Aves_Chinalco_Seca_2018.objects.all().order_by('id').reverse()
    lista_fotos_aves=Fotos_Integrado_Aves_Chinalco_Seca_2018.objects.all().order_by('id').reverse()
    lista_coordenadas_aves=CoordenadasIntegradasAves2018.objects.all().order_by('id').reverse()


    return render(request, 'biologia/lista_formularios/Lista_forms.html',
                  {'lista_aves_chinalco':lista_aves_chinalco,
                   'lista_fotos_aves':lista_fotos_aves,
                   'lista_coordenadas_aves':lista_coordenadas_aves})

# =============== REPTILES ===============
def list_forms_Reptiles_Amfibios(request):
    list_bd_anfibios_reptiles = BD_anfibios_reptiles_BD.objects.all().order_by('id').reverse()
    list_bd_fotos = BD_fotos_BD.objects.all().order_by('id').reverse()
    list_coordenadas = Coordenadas_BD.objects.all().order_by('id').reverse()
    list_hoja1 = hoja1_BD.objects.all()

    return render(request, 'biologia/lista_formularios/lista_forms_anfibios_rectiles.html',
                        {'lista_bd_anfibios_reptiles':list_bd_anfibios_reptiles,
                        'lista_bd_fotos':list_bd_fotos,
                        'lista_coordenadas': list_coordenadas,
                        'lista_hoja1': list_hoja1})

# ================== MAMIFEROS ====================
def list_forms_Mamiferos(request):
    lista_bitacora_mamiferos = bitacora_BD_Mamiferos.objects.all().order_by('id').reverse()
    lista_esfuerzo_mamiferos = Esfuerzo_Muestreo_BD_Mamiferos.objects.all().order_by('id').reverse()
    lista_coordenadas_bd_mamiferos = Coordenadas_reporte_BD_Mamiferos.objects.all().order_by('id').reverse()
    lista_basedatos_mamimeferos = BD_mamiferos_BD_Mamiferos.objects.all().order_by('id').reverse()


    return render(request, 'biologia/lista_formularios/lista_forms_mamiferos.html',
                    {'list_bitacora_mamiferos':lista_bitacora_mamiferos,
                    'list_coordenadas_mamiferos':lista_coordenadas_bd_mamiferos,
                    'list_esfuerzo_mamiferos':lista_esfuerzo_mamiferos,
                    'list_basedatos_mamimeferos':lista_basedatos_mamimeferos})

# ===================== VEGETACION =====================
def list_forms_Vegetacion(request):

    lista_bd_vegetacion = Base_Datos_Vegetacion_2018.objects.all().order_by('id').reverse()
    lista_hoja3_vegetacion = Hoja3_BD_Vegetacion_2018.objects.all().order_by('id').reverse()
    lista_esfuerzo_vegetacion = Esfuerzo_Muestreo_BD_Vegetacion_2018.objects.all().order_by('id').reverse()
    lista_compensacion_vegetacion = Compensacion_Ambiental_Vegetacion_2018.objects.all().order_by('id').reverse()

    return render(request, 'biologia/lista_formularios/lista_forms_vegetacion.html',
                    {'lista_hoja3_vegetacion': lista_hoja3_vegetacion,
                     'lista_bd_vegetacion': lista_bd_vegetacion,
                     'lista_esfuerzo_vegetacion': lista_esfuerzo_vegetacion,
                     'lista_compensacion_vegetacion': lista_compensacion_vegetacion})

#-----------------------------------------------------------------------------------------------------------------------
#Generacion de archivos EXCEL, tabla/Rama/Proyecto
from django.shortcuts import render
from biologia.models import *
from openpyxl import Workbook
from django.http.response import HttpResponse
from openpyxl.styles import Font
from openpyxl.styles import numbers
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
#ant-------
from django.views.generic.base import TemplateView
import xlwt

# ================ SECCION PAJAROS ==================

# BD_AVES_CHINALCO
def export_aves_xls(request):
    coordenadas = BD_Integrada_Aves_Chinalco_Seca_2018.objects.all()

    wb = Workbook()
    ws = wb.active

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble
    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    bor = Border(top=double, left=thin, right=thin, bottom=double)
    fillfg = PatternFill("solid", fgColor="c2d69b")
    fil = fill = GradientFill(stop=("000000", "FFFFFF"))
    font_cell = Font(b=True, color="000000")
    aling = Alignment(horizontal="center", vertical="center")


    #ws['A1'] = 'Usuario'
    ws['A1'] = 'Código de proyecto'
    ws['B1'] = 'Fecha de registro'
    ws['C1'] = 'Temporada de Evaluación'
    ws['D1'] = 'Unidad de vegetación '
    ws['E1'] = 'Estación de muestreo'
    ws['F1'] = 'Unidad de muestreo'
    ws['G1'] = 'Metodo de Muestreo'
    ws['H1'] = 'Periodo del Dia'
    ws['I1'] = 'Hora de Registro'
    ws['J1'] = 'Orden'
    ws['K1'] = 'Familia'
    ws['L1'] = 'Especie'
    ws['M1'] = 'Nombre Comun'
    ws['N1'] = 'Fuente de clasificación taxonomica'
    ws['O1'] = 'X'
    ws['P1'] = 'Y'
    ws['Q1'] = 'Altitud (MSNM)'
    ws['R1'] = 'Datum'
    ws['S1'] = 'Temperie'
    ws['T1'] = 'Tipo de Evaluacion'
    ws['U1'] = 'Registro Visual'
    ws['V1'] = 'Registro Auditivo'
    ws['W1'] = 'Registro por Captura'
    ws['X1'] = 'Registro Casual'
    ws['Y1'] = 'Nº total de individuos'
    ws['Z1'] = 'Distancia al punto de conteo (m)'
    ws['AA1'] = 'Observaciones (etología)'
    ws['AB1'] = 'Peso(gr)'
    ws['AC1'] = 'Sexo'
    ws['AD1'] = 'Longitud(cm)'
    ws['AE1'] = 'Especie segun Ambiente'
    ws['AF1'] = 'Edad Reproductiva'
    ws['AG1'] = 'Conducta Reproductiva'
    ws['AH1'] = 'Tipo de Locomocion'
    ws['AI1'] = 'Tipo de Habitad'
    ws['AJ1'] = 'Grupo Trófico'
    ws['AK1'] = 'Especie Endémica'
    ws['AL1'] = 'Zona de Endemismo'
    ws['AM1'] = 'Restringida a EBA'
    ws['AN1'] = 'Restringida a Bioma (Stozt et al., 1996)'
    ws['AO1'] = 'Categoría de abundancia'
    ws['AP1'] = 'DS O04-2014-MINAGRI'
    ws['AQ1'] = 'IUCN'
    ws['AR1'] = 'CITES'
    ws['AS1'] = 'Migratoria'
    ws['AT1'] = 'Congregatoria'
    ws['AU1'] = 'Observaciones sobre la muestra '
    ws['AV1'] = 'Actividad humana'
    ws['AW1'] = 'Estado de conservación del hábitat'
    ws['AX1'] = 'Colector de Registro'
    ws['AY1'] = 'Tipo de Muestra'
    ws['AZ1'] = 'Estado de Muestra'
    ws['BA1'] = 'Tipo de preparación de la muestra'
    ws['BB1'] = 'Codigo de Colecta'
    ws['BC1'] = 'Institución de depósito'
    ws['BD1'] = 'Código fotos'



    abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
           'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ',
           'BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR','BS','BT','BU','BV','BW','BX','BY','BZ']

    for col in range (1,57):
        cd = ws.cell(row=1,column=col)
        cd.border = bor
        cd.fill = fillfg
        cd.font  = font_cell
        cd.alignment = aling
        ws.column_dimensions[abc[col-1]].width=20.0

    for col in range(1, 57):
        for row in range(1,len(coordenadas)+1):
            cd = ws.cell(row=row+1,column=col)
            cd.border = bor
    cont=2


    for coordenada in coordenadas:
        ws.cell(row=cont,column=1).value = coordenada.codigo_proyecto
        ws.cell(row=cont,column=2,value = coordenada.fecha_registro).number_format = numbers.FORMAT_DATE_XLSX15
        ws.cell(row=cont,column=3).value = coordenada.temporada_evaluacion
        ws.cell(row=cont,column=4).value = coordenada.unidad_vegetacion
        ws.cell(row=cont,column=5).value = coordenada.estacion_muestreo
        ws.cell(row=cont,column=6).value = coordenada.unidad_muestreo
        ws.cell(row=cont,column=7).value = coordenada.metodo_muestreo
        ws.cell(row=cont,column=8).value = coordenada.periodo_dia
        ws.cell(row=cont,column=9).value = coordenada.hora_registro
        ws.cell(row=cont,column=10).value = coordenada.orden
        ws.cell(row=cont,column=11).value = coordenada.familia
        ws.cell(row=cont,column=12).value = coordenada.especie
        ws.cell(row=cont,column=13).value = coordenada.nombre_comun
        ws.cell(row=cont,column=14).value = coordenada.fuente_clasificacion_taxonomica
        ws.cell(row=cont,column=15).value = coordenada.x
        ws.cell(row=cont,column=16).value = coordenada.y
        ws.cell(row=cont,column=17).value = coordenada.altitud_msnm
        ws.cell(row=cont,column=18).value = coordenada.datum
        ws.cell(row=cont,column=19).value = coordenada.temperie
        ws.cell(row=cont,column=20).value = coordenada.tipo_evaluacion
        ws.cell(row=cont,column=21).value = coordenada.registro_visual
        ws.cell(row=cont,column=22).value = coordenada.registro_auditivo
        ws.cell(row=cont,column=23).value = coordenada.registro_captura
        ws.cell(row=cont,column=24).value = coordenada.registro_casual
        ws.cell(row=cont,column=25).value = coordenada.numero_total_individuos
        ws.cell(row=cont,column=26).value = coordenada.distancia_punto_conteo
        ws.cell(row=cont,column=27).value = coordenada.observaciones_etologia
        ws.cell(row=cont,column=28).value = coordenada.peso
        ws.cell(row=cont,column=29).value = coordenada.sexo
        ws.cell(row=cont,column=30).value = coordenada.longitud
        ws.cell(row=cont,column=31).value = coordenada.especie_ambiente
        ws.cell(row=cont,column=32).value = coordenada.edad_reproductiva
        ws.cell(row=cont,column=33).value = coordenada.conducta_reproductiva
        ws.cell(row=cont,column=34).value = coordenada.tipo_locomocion
        ws.cell(row=cont,column=35).value = coordenada.tipo_habitat
        ws.cell(row=cont,column=36).value = coordenada.grupo_triofico
        ws.cell(row=cont,column=37).value = coordenada.especie_endemica
        ws.cell(row=cont,column=38).value = coordenada.zona_endemismo
        ws.cell(row=cont,column=39).value = coordenada.restringida_eba
        ws.cell(row=cont,column=40).value = coordenada.restringida_bioma
        ws.cell(row=cont,column=41).value = coordenada.categoria_abundancia
        ws.cell(row=cont,column=42).value = coordenada.DS_O04_2014_MINAGRI
        ws.cell(row=cont,column=43).value = coordenada.iucn
        ws.cell(row=cont,column=44).value = coordenada.cites
        ws.cell(row=cont,column=45).value = coordenada.migratoria
        ws.cell(row=cont,column=46).value = coordenada.congregatoria
        ws.cell(row=cont,column=47).value = coordenada.observaciones_muestra
        ws.cell(row=cont,column=48).value = coordenada.actividad_humana
        ws.cell(row=cont,column=49).value = coordenada.estado_conservacion_habitat
        ws.cell(row=cont,column=50).value = coordenada.colector_registro
        ws.cell(row=cont,column=51).value = coordenada.tipo_muestra
        ws.cell(row=cont,column=52).value = coordenada.estado_muestra
        ws.cell(row=cont,column=53).value = coordenada.preparacion_muestra
        ws.cell(row=cont,column=54).value = coordenada.codigo_colecta
        ws.cell(row=cont,column=55).value = coordenada.institucion_deposito
        ws.cell(row=cont,column=56).value = coordenada.codigo_fotos


        cont = cont + 1

    nombre_archivo ="BD Aves -Chinalco-seca-2018.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response
#----------------------EXPORT_FOTOS_AVES--------------------------#

def export_fotos_aves_xls(request):
    coordenadas = Fotos_Integrado_Aves_Chinalco_Seca_2018.objects.all()

    wb = Workbook()
    ws = wb.active

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble
    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    bor = Border(top=double, left=thin, right=thin, bottom=double)
    fillfg = PatternFill("solid", fgColor="2a90e6")
    fil = fill = GradientFill(stop=("000000", "FFFFFF"))
    font_cell = Font(b=True, color="000000")
    aling = Alignment(horizontal="center", vertical="center")


    #ws['A1'] = 'Usuario'
    ws['A1'] = 'Estacion de Muestreo'
    ws['B1'] = 'Componente de Proyecto'
    ws['C1'] = 'Unidad de Muestreo'
    ws['D1'] = 'Codigo de la Foto '
    ws['E1'] = 'Fecha de la Toma'
    ws['F1'] = 'Especie'
    ws['G1'] = 'Paisaje'
    ws['H1'] = 'Nombre científico'
    ws['I1'] = 'Nombre común'
    ws['J1'] = 'Nombre local'
    ws['K1'] = 'Distrito'
    ws['L1'] = 'Provincia'
    ws['M1'] = 'Departamento'



    abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M']

    for col in range (1,14):
        cd = ws.cell(row=1,column=col)
        cd.border = bor
        cd.fill = fillfg
        cd.font  = font_cell
        cd.alignment = aling
        ws.column_dimensions[abc[col-1]].width=20.0

    for col in range(1, 14):
        for row in range(1,len(coordenadas)+1):
            cd = ws.cell(row=row+1,column=col)
            cd.border = bor
    cont=2


    for coordenada in coordenadas:
        ws.cell(row=cont,column=1).value = coordenada.estacion_muestreo
        ws.cell(row=cont,column=2).value = coordenada.componente_proyecto
        ws.cell(row=cont,column=3).value = coordenada.unidad_muestreo
        ws.cell(row=cont,column=4).value = coordenada.codigo_foto
        ws.cell(row=cont,column=5,value = coordenada.fecha_toma).number_format = numbers.FORMAT_DATE_XLSX15
        ws.cell(row=cont,column=6).value = coordenada.especie
        ws.cell(row=cont,column=7).value = coordenada.paisaje
        ws.cell(row=cont,column=8).value = coordenada.nombre_cientifica
        ws.cell(row=cont,column=9).value = coordenada.nombre_comun
        ws.cell(row=cont,column=10).value = coordenada.nombre_local
        ws.cell(row=cont,column=11).value = coordenada.distrito
        ws.cell(row=cont,column=12).value = coordenada.provincia
        ws.cell(row=cont,column=13).value = coordenada.departamento

        cont = cont + 1

    nombre_archivo ="Fotos integrado.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response


# COORDENADAS_AVES_CHINALCO
def export_coordenadas_aves_xls(request):
    coordenadas = CoordenadasIntegradasAves2018.objects.all()

    wb = Workbook()
    ws = wb.active

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble
    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    bor = Border(top=double, left=thin, right=thin, bottom=double)
    fillfg = PatternFill("solid", fgColor="f72626")
    fil = fill = GradientFill(stop=("000000", "FFFFFF"))
    font_cell = Font(b=True, color="000000")
    aling = Alignment(horizontal="center", vertical="center")


    #ws['A1'] = 'Usuario'
    ws['A1'] = 'Unidad de vegetación'
    ws['B1'] = 'Estación'
    ws['C1'] = 'Punto de conteo'
    ws['D1'] = 'Zona '
    ws['E1'] = 'Este'
    ws['F1'] = 'Norte'
    ws['G1'] = 'Altitud'
    ws['H1'] = 'Fecha'
    ws['I1'] = 'Temperie'
    ws['J1'] = 'Conservación de la estacion de muestreo '
    ws['K1'] = 'Impacto o actividad humana cercana'
    ws['L1'] = 'Descripcion del Punto de conteo'


    abc = ['A','B','C','D','E','F','G','H','I','J','K','L']

    for col in range (1,13):
        cd = ws.cell(row=1,column=col)
        cd.border = bor
        cd.fill = fillfg
        cd.font  = font_cell
        cd.alignment = aling
        ws.column_dimensions[abc[col-1]].width=20.0

    for col in range(1, 13):
        for row in range(1,len(coordenadas)+1):
            cd = ws.cell(row=row+1,column=col)
            cd.border = bor
    cont=2


    for coordenada in coordenadas:
        ws.cell(row=cont,column=1).value = coordenada.unidad_vegetacion
        ws.cell(row=cont,column=2).value = coordenada.estacion
        ws.cell(row=cont,column=3).value = coordenada.punto_conteo
        ws.cell(row=cont,column=4).value = coordenada.zona
        ws.cell(row=cont,column=5).value = coordenada.este1
        ws.cell(row=cont,column=6).value = coordenada.norte1
        ws.cell(row=cont,column=7).value = coordenada.altitud
        ws.cell(row=cont,column=8,value = coordenada.fecha_registro).number_format = numbers.FORMAT_DATE_XLSX15
        ws.cell(row=cont,column=9).value = coordenada.temperie
        ws.cell(row=cont,column=10).value = coordenada.conservacion_estacion_muestreo
        ws.cell(row=cont,column=11).value = coordenada.impacto_actividad_humana_cercana
        ws.cell(row=cont,column=12).value = coordenada.descripcion_punto_conteo

        cont = cont + 1

    nombre_archivo ="Coordenadas integradas.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

#==============================EXPORTACION UNICA AVES========================#

#----------------------EXPORT_FOTOS_AVES_UNICA--------------------------#

def export_fotos_aves_unica_xls(request,id):
    coordenadas = Fotos_Integrado_Aves_Chinalco_Seca_2018.objects.get(id=id)

    wb = Workbook()
    ws = wb.active

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble
    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    bor = Border(top=double, left=thin, right=thin, bottom=double)
    fillfg = PatternFill("solid", fgColor="2a90e6")
    fil = fill = GradientFill(stop=("000000", "FFFFFF"))
    font_cell = Font(b=True, color="000000")
    aling = Alignment(horizontal="center", vertical="center")


    #ws['A1'] = 'Usuario'
    ws['A1'] = 'Estacion de Muestreo'
    ws['B1'] = 'Componente de Proyecto'
    ws['C1'] = 'Unidad de Muestreo'
    ws['D1'] = 'Codigo de la Foto '
    ws['E1'] = 'Fecha de la Toma'
    ws['F1'] = 'Especie'
    ws['G1'] = 'Paisaje'
    ws['H1'] = 'Nombre científico'
    ws['I1'] = 'Nombre común'
    ws['J1'] = 'Nombre local'
    ws['K1'] = 'Distrito'
    ws['L1'] = 'Provincia'
    ws['M1'] = 'Departamento'



    abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M']

    for col in range (1,14):
        cd = ws.cell(row=1,column=col)
        cd.border = bor
        cd.fill = fillfg
        cd.font  = font_cell
        cd.alignment = aling
        ws.column_dimensions[abc[col-1]].width=20.0

    for col in range(1, 14):
        cd = ws.cell(row=2,column=col)
        cd.border = bor
    cont=2


    ws.cell(row=cont,column=1).value = coordenadas.estacion_muestreo
    ws.cell(row=cont,column=2).value = coordenadas.componente_proyecto
    ws.cell(row=cont,column=3).value = coordenadas.unidad_muestreo
    ws.cell(row=cont,column=4).value = coordenadas.codigo_foto
    ws.cell(row=cont,column=5,value = coordenadas.fecha_toma).number_format = numbers.FORMAT_DATE_XLSX15
    ws.cell(row=cont,column=6).value = coordenadas.especie
    ws.cell(row=cont,column=7).value = coordenadas.paisaje
    ws.cell(row=cont,column=8).value = coordenadas.nombre_cientifica
    ws.cell(row=cont,column=9).value = coordenadas.nombre_comun
    ws.cell(row=cont,column=10).value = coordenadas.nombre_local
    ws.cell(row=cont,column=11).value = coordenadas.distrito
    ws.cell(row=cont,column=12).value = coordenadas.provincia
    ws.cell(row=cont,column=13).value = coordenadas.departamento

    cont = cont + 1

    nombre_archivo ="Fotos integrado Unico.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

# BD_AVES_CHINALCO_UNICA
def export_aves_unica_xls(request,id):
    coordenadas = BD_Integrada_Aves_Chinalco_Seca_2018.objects.get(id=id)

    wb = Workbook()
    ws = wb.active

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble
    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    bor = Border(top=double, left=thin, right=thin, bottom=double)
    fillfg = PatternFill("solid", fgColor="c2d69b")
    fil = fill = GradientFill(stop=("000000", "FFFFFF"))
    font_cell = Font(b=True, color="000000")
    aling = Alignment(horizontal="center", vertical="center")


    #ws['A1'] = 'Usuario'
    ws['A1'] = 'Código de proyecto'
    ws['B1'] = 'Fecha de registro'
    ws['C1'] = 'Temporada de Evaluación'
    ws['D1'] = 'Unidad de vegetación '
    ws['E1'] = 'Estación de muestreo'
    ws['F1'] = 'Unidad de muestreo'
    ws['G1'] = 'Metodo de Muestreo'
    ws['H1'] = 'Periodo del Dia'
    ws['I1'] = 'Hora de Registro'
    ws['J1'] = 'Orden'
    ws['K1'] = 'Familia'
    ws['L1'] = 'Especie'
    ws['M1'] = 'Nombre Comun'
    ws['N1'] = 'Fuente de clasificación taxonomica'
    ws['O1'] = 'X'
    ws['P1'] = 'Y'
    ws['Q1'] = 'Altitud (MSNM)'
    ws['R1'] = 'Datum'
    ws['S1'] = 'Temperie'
    ws['T1'] = 'Tipo de Evaluacion'
    ws['U1'] = 'Registro Visual'
    ws['V1'] = 'Registro Auditivo'
    ws['W1'] = 'Registro por Captura'
    ws['X1'] = 'Registro Casual'
    ws['Y1'] = 'Nº total de individuos'
    ws['Z1'] = 'Distancia al punto de conteo (m)'
    ws['AA1'] = 'Observaciones (etología)'
    ws['AB1'] = 'Peso(gr)'
    ws['AC1'] = 'Sexo'
    ws['AD1'] = 'Longitud(cm)'
    ws['AE1'] = 'Especie segun Ambiente'
    ws['AF1'] = 'Edad Reproductiva'
    ws['AG1'] = 'Conducta Reproductiva'
    ws['AH1'] = 'Tipo de Locomocion'
    ws['AI1'] = 'Tipo de Habitad'
    ws['AJ1'] = 'Grupo Trófico'
    ws['AK1'] = 'Especie Endémica'
    ws['AL1'] = 'Zona de Endemismo'
    ws['AM1'] = 'Restringida a EBA'
    ws['AN1'] = 'Restringida a Bioma (Stozt et al., 1996)'
    ws['AO1'] = 'Categoría de abundancia'
    ws['AP1'] = 'DS O04-2014-MINAGRI'
    ws['AQ1'] = 'IUCN'
    ws['AR1'] = 'CITES'
    ws['AS1'] = 'Migratoria'
    ws['AT1'] = 'Congregatoria'
    ws['AU1'] = 'Observaciones sobre la muestra '
    ws['AV1'] = 'Actividad humana'
    ws['AW1'] = 'Estado de conservación del hábitat'
    ws['AX1'] = 'Colector de Registro'
    ws['AY1'] = 'Tipo de Muestra'
    ws['AZ1'] = 'Estado de Muestra'
    ws['BA1'] = 'Tipo de preparación de la muestra'
    ws['BB1'] = 'Codigo de Colecta'
    ws['BC1'] = 'Institución de depósito'
    ws['BD1'] = 'Código fotos'



    abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
           'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ',
           'BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR','BS','BT','BU','BV','BW','BX','BY','BZ']

    for col in range (1,57):
        cd = ws.cell(row=1,column=col)
        cd.border = bor
        cd.fill = fillfg
        cd.font  = font_cell
        cd.alignment = aling
        ws.column_dimensions[abc[col-1]].width=20.0

    for col in range(1, 57):
        cd = ws.cell(row=2,column=col)
        cd.border = bor
    cont=2



    ws.cell(row=cont,column=1).value = coordenadas.codigo_proyecto
    ws.cell(row=cont,column=2,value = coordenadas.fecha_registro).number_format = numbers.FORMAT_DATE_XLSX15
    ws.cell(row=cont,column=3).value = coordenadas.temporada_evaluacion
    ws.cell(row=cont,column=4).value = coordenadas.unidad_vegetacion
    ws.cell(row=cont,column=5).value = coordenadas.estacion_muestreo
    ws.cell(row=cont,column=6).value = coordenadas.unidad_muestreo
    ws.cell(row=cont,column=7).value = coordenadas.metodo_muestreo
    ws.cell(row=cont,column=8).value = coordenadas.periodo_dia
    ws.cell(row=cont,column=9).value = coordenadas.hora_registro
    ws.cell(row=cont,column=10).value = coordenadas.orden
    ws.cell(row=cont,column=11).value = coordenadas.familia
    ws.cell(row=cont,column=12).value = coordenadas.especie
    ws.cell(row=cont,column=13).value = coordenadas.nombre_comun
    ws.cell(row=cont,column=14).value = coordenadas.fuente_clasificacion_taxonomica
    ws.cell(row=cont,column=15).value = coordenadas.x
    ws.cell(row=cont,column=16).value = coordenadas.y
    ws.cell(row=cont,column=17).value = coordenadas.altitud_msnm
    ws.cell(row=cont,column=18).value = coordenadas.datum
    ws.cell(row=cont,column=19).value = coordenadas.temperie
    ws.cell(row=cont,column=20).value = coordenadas.tipo_evaluacion
    ws.cell(row=cont,column=21).value = coordenadas.registro_visual
    ws.cell(row=cont,column=22).value = coordenadas.registro_auditivo
    ws.cell(row=cont,column=23).value = coordenadas.registro_captura
    ws.cell(row=cont,column=24).value = coordenadas.registro_casual
    ws.cell(row=cont,column=25).value = coordenadas.numero_total_individuos
    ws.cell(row=cont,column=26).value = coordenadas.distancia_punto_conteo
    ws.cell(row=cont,column=27).value = coordenadas.observaciones_etologia
    ws.cell(row=cont,column=28).value = coordenadas.peso
    ws.cell(row=cont,column=29).value = coordenadas.sexo
    ws.cell(row=cont,column=30).value = coordenadas.longitud
    ws.cell(row=cont,column=31).value = coordenadas.especie_ambiente
    ws.cell(row=cont,column=32).value = coordenadas.edad_reproductiva
    ws.cell(row=cont,column=33).value = coordenadas.conducta_reproductiva
    ws.cell(row=cont,column=34).value = coordenadas.tipo_locomocion
    ws.cell(row=cont,column=35).value = coordenadas.tipo_habitat
    ws.cell(row=cont,column=36).value = coordenadas.grupo_triofico
    ws.cell(row=cont,column=37).value = coordenadas.especie_endemica
    ws.cell(row=cont,column=38).value = coordenadas.zona_endemismo
    ws.cell(row=cont,column=39).value = coordenadas.restringida_eba
    ws.cell(row=cont,column=40).value = coordenadas.restringida_bioma
    ws.cell(row=cont,column=41).value = coordenadas.categoria_abundancia
    ws.cell(row=cont,column=42).value = coordenadas.DS_O04_2014_MINAGRI
    ws.cell(row=cont,column=43).value = coordenadas.iucn
    ws.cell(row=cont,column=44).value = coordenadas.cites
    ws.cell(row=cont,column=45).value = coordenadas.migratoria
    ws.cell(row=cont,column=46).value = coordenadas.congregatoria
    ws.cell(row=cont,column=47).value = coordenadas.observaciones_muestra
    ws.cell(row=cont,column=48).value = coordenadas.actividad_humana
    ws.cell(row=cont,column=49).value = coordenadas.estado_conservacion_habitat
    ws.cell(row=cont,column=50).value = coordenadas.colector_registro
    ws.cell(row=cont,column=51).value = coordenadas.tipo_muestra
    ws.cell(row=cont,column=52).value = coordenadas.estado_muestra
    ws.cell(row=cont,column=53).value = coordenadas.preparacion_muestra
    ws.cell(row=cont,column=54).value = coordenadas.codigo_colecta
    ws.cell(row=cont,column=55).value = coordenadas.institucion_deposito
    ws.cell(row=cont,column=56).value = coordenadas.codigo_fotos

    cont = cont + 1

    nombre_archivo ="BD Aves -Chinalco-seca-2018_Unico.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

# COORDENADAS_AVES_CHINALCO_UNICO
def export_coordenadas_aves_unica_xls(request,id):
    coordenadas = CoordenadasIntegradasAves2018.objects.get(id=id)

    wb = Workbook()
    ws = wb.active

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble
    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    bor = Border(top=double, left=thin, right=thin, bottom=double)
    fillfg = PatternFill("solid", fgColor="f72626")
    fil = fill = GradientFill(stop=("000000", "FFFFFF"))
    font_cell = Font(b=True, color="000000")
    aling = Alignment(horizontal="center", vertical="center")


    #ws['A1'] = 'Usuario'
    ws['A1'] = 'Unidad de vegetación'
    ws['B1'] = 'Estación'
    ws['C1'] = 'Punto de conteo'
    ws['D1'] = 'Zona '
    ws['E1'] = 'Este'
    ws['F1'] = 'Norte'
    ws['G1'] = 'Altitud'
    ws['H1'] = 'Fecha'
    ws['I1'] = 'Temperie'
    ws['J1'] = 'Conservación de la estacion de muestreo '
    ws['K1'] = 'Impacto o actividad humana cercana'
    ws['L1'] = 'Descripcion del Punto de conteo'


    abc = ['A','B','C','D','E','F','G','H','I','J','K','L']

    for col in range (1,13):
        cd = ws.cell(row=1,column=col)
        cd.border = bor
        cd.fill = fillfg
        cd.font  = font_cell
        cd.alignment = aling
        ws.column_dimensions[abc[col-1]].width=20.0

    for col in range(1, 13):
        cd = ws.cell(row=2,column=col)
        cd.border = bor
    cont=2


    ws.cell(row=cont,column=1).value = coordenadas.unidad_vegetacion
    ws.cell(row=cont,column=2).value = coordenadas.estacion
    ws.cell(row=cont,column=3).value = coordenadas.punto_conteo
    ws.cell(row=cont,column=4).value = coordenadas.zona
    ws.cell(row=cont,column=5).value = coordenadas.este1
    ws.cell(row=cont,column=6).value = coordenadas.norte1
    ws.cell(row=cont,column=7).value = coordenadas.altitud
    ws.cell(row=cont,column=8,value = coordenadas.fecha_registro).number_format = numbers.FORMAT_DATE_XLSX15
    ws.cell(row=cont,column=9).value = coordenadas.temperie
    ws.cell(row=cont,column=10).value = coordenadas.conservacion_estacion_muestreo
    ws.cell(row=cont,column=11).value = coordenadas.impacto_actividad_humana_cercana
    ws.cell(row=cont,column=12).value = coordenadas.descripcion_punto_conteo

    cont = cont + 1

    nombre_archivo ="Coordenadas integradas Unico.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

# ============================== SECCION REPTILES ======================================

##----------------------EXPORT_BD_ANFIBIOS_REPTILES-------------------------#

def export_Anfibios_reptiles_BD_xls(request):

    BDReptilesAnfibios = BD_anfibios_reptiles_BD.objects.all()

    wb = Workbook()
    ws = wb.active

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble
    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    bor = Border(top=double, left=thin, right=thin, bottom=double)
    fillfg = PatternFill("solid", fgColor="f72626")
    fil = fill = GradientFill(stop=("000000", "FFFFFF"))
    font_cell = Font(b=True, color="000000")
    aling = Alignment(horizontal="center", vertical="center")

    ws['A1'] = 'Codigo de proyecto'
    ws['B1'] = 'Fecha de registro'
    ws['C1'] = 'Temporada de evaluacion'
    ws['D1'] = 'Unidad de vegetación (Quebrada)'
    ws['E1'] = 'Sector'
    ws['F1'] = 'Codigo de la quebrada'
    ws['G1'] = 'Disciplina biologia'
    ws['H1'] = 'Estacion de muestreo'
    ws['I1'] = 'Unidad de muestreo'
    ws['J1'] = 'Metodo de muestreo'
    ws['K1'] = 'Tamaño de la unidad de muestreo'
    ws['L1'] = '#Chip/marca dedos'
    ws['M1'] = 'Periodo del dia'
    ws['N1'] = 'Hora de inicia de evaluación'
    ws['O1'] = 'Hora de finalización de evaluacion'
    ws['P1'] = 'Hora de registro'
    ws['Q1'] = 'Familia'
    ws['R1'] = 'Especie'
    ws['S1'] = 'Nombre común'
    ws['T1'] = 'Nombre local'
    ws['U1'] = 'Fuente de clasificación taxonomica'
    ws['V1'] = 'X'
    ws['W1'] = 'Y'
    ws['X1'] = 'Proyección'
    ws['Y1'] = 'Altitud (msnm)'
    ws['Z1'] = 'Datum'
    ws['AA1'] = 'T° del ambiente(°C)'
    ws['AB1'] = 'T° del suelo(°C)'
    ws['AC1'] = 'Profundidad de horjarasca(cm)'
    ws['AD1'] = 'Temperie'
    ws['AE1'] = 'Distancia perpendicular de la unidad de muestreo (cm)'
    ws['AF1'] = 'Altura del registro (Distancia vertical al suelo)cm'
    ws['AG1'] = 'Tipo de evaluacion'
    ws['AH1'] = 'Tipo de registro'
    ws['AI1'] = 'N° Registro directo'
    ws['AJ1'] = 'N° Registro indirecto'
    ws['AK1'] = 'N° Registro por captura'
    ws['AL1'] = 'Estado del hallazgo'
    ws['AM1'] = 'Etología (Comportamiento del individuo registrado)'
    ws['AN1'] = 'Nº total de individuos'
    ws['AO1'] = 'Peso (gr)'
    ws['AP1'] = 'Sexo'
    ws['AQ1'] = 'Longitud total (cm)'
    ws['AR1'] = 'Longitud ano - hocico (cm)'
    ws['AS1'] = 'Longitud cola (lagartijas y serpientes)'
    ws['AT1'] = 'Tipo de Reproducción (Anfibios)'
    ws['AU1'] = 'Conducta reproductiva (in situ)'
    ws['AV1'] = 'Conducta Social'
    ws['AW1'] = 'Territorialidad'
    ws['AX1'] = 'Hábito'
    ws['AY1'] = 'Tipo de Hábitat (in situ)l'
    ws['AZ1'] = 'Micro-habitat (insitu)'
    ws['BA1'] = 'Plantas dominante del microhabitat (insitu)'
    ws['BB1'] = 'Grupo Trófico'
    ws['BC1'] = 'Registro de dieta  (in situ)'
    ws['BD1'] = 'Especie Endémica'
    ws['BE1'] = 'Nivel de Endemismo'
    ws['BF1'] = 'Distribución de especies endémicas'
    ws['BG1'] = 'DS O34-2004-AG'
    ws['BH1'] = 'IUCN'
    ws['BI1'] = 'CITES'
    ws['BJ1'] = 'Uso potencial'
    ws['BK1'] = 'Observaciones del registro'
    ws['BL1'] = 'Actividad humana'
    ws['BM1'] = 'Estado de conservación del hábitat'
    ws['BN1'] = 'Distancia al componente minero'
    ws['BO1'] = 'Medida de control para impedir o mitigar ingreso de grupo biologico(cercas, espantapajaros, birdball)'
    ws['BP1'] = 'Micro-habitat'
    ws['BQ1'] = 'Animal a las instalaciones (cercas rotas, de cocada amplia, espantapajaro en mal estado)'
    ws['BR1'] = 'Riesgo de contacto por el animal a las instalaciones (Presencia de vegetación ribereña, Matorrales, Floración, Fructificación, Pastizales vigorosos, Ojos de agua, Roquedales '
    ws['BS1'] = 'Observaciones (animales muertos)'

    abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
               'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ',
               'BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR','BS','BT','BU','BV','BW','BX','BY','BZ']

    for col in range (1,72):
        cd = ws.cell(row=1,column=col)
        cd.border = bor
        cd.fill = fillfg
        cd.font  = font_cell
        cd.alignment = aling
        ws.column_dimensions[abc[col-1]].width=20.0

    for col in range(1, 72):
        for row in range(1,len(BDReptilesAnfibios)+1):
            cd = ws.cell(row=row+2,column=col)
            cd.border = bor

    cont=2

    for Bd_ANFIBIOSREPTILES in BDReptilesAnfibios:
        #ws.cell(row=cont,column=1).value = Bd_ANFIBIOSREPTILES.user
        #ws.cell(row=cont,column=1).value = Bd_ANFIBIOSREPTILES.codigo_proyecto
        ws.cell(row=cont,column=1).value = Bd_ANFIBIOSREPTILES.codigo_proyecto
        ws.cell(row=cont,column=2).value = Bd_ANFIBIOSREPTILES.fecha_registro
        ws.cell(row=cont,column=3).value = Bd_ANFIBIOSREPTILES.temporada_evaluacion
        ws.cell(row=cont,column=4).value = Bd_ANFIBIOSREPTILES.unidad_vegetacion_quebrada
        ws.cell(row=cont,column=5).value = Bd_ANFIBIOSREPTILES.sector
        ws.cell(row=cont,column=6).value = Bd_ANFIBIOSREPTILES.codigo_quebrada
        ws.cell(row=cont,column=7).value = Bd_ANFIBIOSREPTILES.disciplina_biologica
        ws.cell(row=cont,column=8).value = Bd_ANFIBIOSREPTILES.estacion_muestreo
        ws.cell(row=cont,column=9).value = Bd_ANFIBIOSREPTILES.unidad_muestreo
        ws.cell(row=cont,column=10).value = Bd_ANFIBIOSREPTILES.metodo_muestreo
        ws.cell(row=cont,column=11).value = Bd_ANFIBIOSREPTILES.tamaño_unidad_muestreo
        ws.cell(row=cont,column=12).value = Bd_ANFIBIOSREPTILES.chip_marca_dedos
        ws.cell(row=cont,column=13).value = Bd_ANFIBIOSREPTILES.periodo_dia
        ws.cell(row=cont,column=14).value = Bd_ANFIBIOSREPTILES.hora_inicio_evaluacion
        ws.cell(row=cont,column=15).value = Bd_ANFIBIOSREPTILES.hora_finalizacion_evaluacion
        ws.cell(row=cont,column=16).value = Bd_ANFIBIOSREPTILES.hora_registro
        ws.cell(row=cont,column=17).value = Bd_ANFIBIOSREPTILES.familia
        ws.cell(row=cont,column=18).value = Bd_ANFIBIOSREPTILES.especie
        ws.cell(row=cont,column=19).value = Bd_ANFIBIOSREPTILES.nombre_comun
        ws.cell(row=cont,column=20).value = Bd_ANFIBIOSREPTILES.nombre_local
        ws.cell(row=cont,column=21).value = Bd_ANFIBIOSREPTILES.fuente_clasificacion_taxonomica
        ws.cell(row=cont,column=22).value = Bd_ANFIBIOSREPTILES.x
        ws.cell(row=cont,column=23).value = Bd_ANFIBIOSREPTILES.y
        ws.cell(row=cont,column=24).value = Bd_ANFIBIOSREPTILES.proyeccion
        ws.cell(row=cont,column=25).value = Bd_ANFIBIOSREPTILES.altitud_msnm
        ws.cell(row=cont,column=26).value = Bd_ANFIBIOSREPTILES.datum
        ws.cell(row=cont,column=27).value = Bd_ANFIBIOSREPTILES.temperatura_ambiente
        ws.cell(row=cont,column=28).value = Bd_ANFIBIOSREPTILES.temperatura_suelo
        ws.cell(row=cont,column=29).value = Bd_ANFIBIOSREPTILES.profundidad_hojarrasca_cm
        ws.cell(row=cont,column=30).value = Bd_ANFIBIOSREPTILES.temperie
        ws.cell(row=cont,column=31).value = Bd_ANFIBIOSREPTILES.distancia_perpendicular_unidad_muestreo
        ws.cell(row=cont,column=32).value = Bd_ANFIBIOSREPTILES.altura_registro_distancia_vertical_suelo
        ws.cell(row=cont,column=33).value = Bd_ANFIBIOSREPTILES.tipo_evaluacion
        ws.cell(row=cont,column=34).value = Bd_ANFIBIOSREPTILES.tipo_registro
        ws.cell(row=cont,column=35).value = Bd_ANFIBIOSREPTILES.numero_registro_directo
        ws.cell(row=cont,column=36).value = Bd_ANFIBIOSREPTILES.numero_registro_indirecto
        ws.cell(row=cont,column=37).value = Bd_ANFIBIOSREPTILES.numero_registro_captura
        ws.cell(row=cont,column=38).value = Bd_ANFIBIOSREPTILES.estado_hallazgo
        ws.cell(row=cont,column=39).value = Bd_ANFIBIOSREPTILES.etologia_comportamiento_individuo_registrado
        ws.cell(row=cont,column=40).value = Bd_ANFIBIOSREPTILES.numero_total_individuos
        ws.cell(row=cont,column=41).value = Bd_ANFIBIOSREPTILES.peso
        ws.cell(row=cont,column=42).value = Bd_ANFIBIOSREPTILES.sexo
        ws.cell(row=cont,column=43).value = Bd_ANFIBIOSREPTILES.longitud
        ws.cell(row=cont,column=44).value = Bd_ANFIBIOSREPTILES.longitud_ano_hocico
        ws.cell(row=cont,column=45).value = Bd_ANFIBIOSREPTILES.longitud_cola_lagartijas_serpientes
        ws.cell(row=cont,column=46).value = Bd_ANFIBIOSREPTILES.tipo_reproduccion_anfibios
        ws.cell(row=cont,column=47).value = Bd_ANFIBIOSREPTILES.conducta_reproductva_in_situ
        ws.cell(row=cont,column=48).value = Bd_ANFIBIOSREPTILES.conducta_social
        ws.cell(row=cont,column=49).value = Bd_ANFIBIOSREPTILES.territorialidad
        ws.cell(row=cont,column=50).value = Bd_ANFIBIOSREPTILES.habito
        ws.cell(row=cont,column=51).value = Bd_ANFIBIOSREPTILES.tipo_habitat_in_situ
        ws.cell(row=cont,column=52).value = Bd_ANFIBIOSREPTILES.micro_habitat_in_situ
        ws.cell(row=cont,column=53).value = Bd_ANFIBIOSREPTILES.planta_dominante_microhabitat_in_situ
        ws.cell(row=cont,column=54).value = Bd_ANFIBIOSREPTILES.grupo_trofico
        ws.cell(row=cont,column=55).value = Bd_ANFIBIOSREPTILES.registro_dieta_in_situ
        ws.cell(row=cont,column=56).value = Bd_ANFIBIOSREPTILES.especie_endemica
        ws.cell(row=cont,column=57).value = Bd_ANFIBIOSREPTILES.nivel_endemismo
        ws.cell(row=cont,column=58).value = Bd_ANFIBIOSREPTILES.distribucion_especie_endemica
        ws.cell(row=cont,column=59).value = Bd_ANFIBIOSREPTILES.DSO342004AG
        ws.cell(row=cont,column=60).value = Bd_ANFIBIOSREPTILES.IUCN
        ws.cell(row=cont,column=61).value = Bd_ANFIBIOSREPTILES.CITES
        ws.cell(row=cont,column=62).value = Bd_ANFIBIOSREPTILES.uso_potencial
        ws.cell(row=cont,column=63).value = Bd_ANFIBIOSREPTILES.observaciones_registro
        ws.cell(row=cont,column=64).value = Bd_ANFIBIOSREPTILES.actividad_humana
        ws.cell(row=cont,column=65).value = Bd_ANFIBIOSREPTILES.estado_conservacion_habitat
        ws.cell(row=cont,column=66).value = Bd_ANFIBIOSREPTILES.distancia_componente_minero
        ws.cell(row=cont,column=67).value = Bd_ANFIBIOSREPTILES.medida_control_impedir_ingreso
        ws.cell(row=cont,column=68).value = Bd_ANFIBIOSREPTILES.micro_habitat
        ws.cell(row=cont,column=69).value = Bd_ANFIBIOSREPTILES.animal_instalaciones
        ws.cell(row=cont,column=70).value = Bd_ANFIBIOSREPTILES.riesgo_contacto_animal_instalaciones
        ws.cell(row=cont,column=71).value = Bd_ANFIBIOSREPTILES.observaciones_animales_muertos
        cont = cont + 1

    nombre_archivo ="Anfibios_reptiles_BD_2018.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

#DB_Reptiles_unico
def export_Anfibios_reptiles_BD_unico(request, id):

    Bd_ANFIBIOSREPTILES = BD_anfibios_reptiles_BD.objects.get(id=id)

    wb = Workbook()
    ws = wb.active

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble
    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    bor = Border(top=double, left=thin, right=thin, bottom=double)
    fillfg = PatternFill("solid", fgColor="f72626")
    fil = fill = GradientFill(stop=("000000", "FFFFFF"))
    font_cell = Font(b=True, color="000000")
    aling = Alignment(horizontal="center", vertical="center")

    ws['A1'] = 'Codigo de proyecto'
    ws['B1'] = 'Fecha de registro'
    ws['C1'] = 'Temporada de evaluacion'
    ws['D1'] = 'Unidad de vegetación (Quebrada)'
    ws['E1'] = 'Sector'
    ws['F1'] = 'Codigo de la quebrada'
    ws['G1'] = 'Disciplina biologia'
    ws['H1'] = 'Estacion de muestreo'
    ws['I1'] = 'Unidad de muestreo'
    ws['J1'] = 'Metodo de muestreo'
    ws['K1'] = 'Tamaño de la unidad de muestreo'
    ws['L1'] = '#Chip/marca dedos'
    ws['M1'] = 'Periodo del dia'
    ws['N1'] = 'Hora de inicia de evaluación'
    ws['O1'] = 'Hora de finalización de evaluacion'
    ws['P1'] = 'Hora de registro'
    ws['Q1'] = 'Familia'
    ws['R1'] = 'Especie'
    ws['S1'] = 'Nombre común'
    ws['T1'] = 'Nombre local'
    ws['U1'] = 'Fuente de clasificación taxonomica'
    ws['V1'] = 'X'
    ws['W1'] = 'Y'
    ws['X1'] = 'Proyección'
    ws['Y1'] = 'Altitud (msnm)'
    ws['Z1'] = 'Datum'
    ws['AA1'] = 'T° del ambiente(°C)'
    ws['AB1'] = 'T° del suelo(°C)'
    ws['AC1'] = 'Profundidad de horjarasca(cm)'
    ws['AD1'] = 'Temperie'
    ws['AE1'] = 'Distancia perpendicular de la unidad de muestreo (cm)'
    ws['AF1'] = 'Altura del registro (Distancia vertical al suelo)cm'
    ws['AG1'] = 'Tipo de evaluacion'
    ws['AH1'] = 'Tipo de registro'
    ws['AI1'] = 'N° Registro directo'
    ws['AJ1'] = 'N° Registro indirecto'
    ws['AK1'] = 'N° Registro por captura'
    ws['AL1'] = 'Estado del hallazgo'
    ws['AM1'] = 'Etología (Comportamiento del individuo registrado)'
    ws['AN1'] = 'Nº total de individuos'
    ws['AO1'] = 'Peso (gr)'
    ws['AP1'] = 'Sexo'
    ws['AQ1'] = 'Longitud total (cm)'
    ws['AR1'] = 'Longitud ano - hocico (cm)'
    ws['AS1'] = 'Longitud cola (lagartijas y serpientes)'
    ws['AT1'] = 'Tipo de Reproducción (Anfibios)'
    ws['AU1'] = 'Conducta reproductiva (in situ)'
    ws['AV1'] = 'Conducta Social'
    ws['AW1'] = 'Territorialidad'
    ws['AX1'] = 'Hábito'
    ws['AY1'] = 'Tipo de Hábitat (in situ)l'
    ws['AZ1'] = 'Micro-habitat (insitu)'
    ws['BA1'] = 'Plantas dominante del microhabitat (insitu)'
    ws['BB1'] = 'Grupo Trófico'
    ws['BC1'] = 'Registro de dieta  (in situ)'
    ws['BD1'] = 'Especie Endémica'
    ws['BE1'] = 'Nivel de Endemismo'
    ws['BF1'] = 'Distribución de especies endémicas'
    ws['BG1'] = 'DS O34-2004-AG'
    ws['BH1'] = 'IUCN'
    ws['BI1'] = 'CITES'
    ws['BJ1'] = 'Uso potencial'
    ws['BK1'] = 'Observaciones del registro'
    ws['BL1'] = 'Actividad humana'
    ws['BM1'] = 'Estado de conservación del hábitat'
    ws['BN1'] = 'Distancia al componente minero'
    ws['BO1'] = 'Medida de control para impedir o mitigar ingreso de grupo biologico(cercas, espantapajaros, birdball)'
    ws['BP1'] = 'Micro-habitat'
    ws['BQ1'] = 'Animal a las instalaciones (cercas rotas, de cocada amplia, espantapajaro en mal estado)'
    ws['BR1'] = 'Riesgo de contacto por el animal a las instalaciones (Presencia de vegetación ribereña, Matorrales, Floración, Fructificación, Pastizales vigorosos, Ojos de agua, Roquedales '
    ws['BS1'] = 'Observaciones (animales muertos)'

    cont=2

    abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
               'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ',
               'BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR','BS','BT','BU','BV','BW','BX','BY','BZ']

    for col in range (1,72):
        cd = ws.cell(row=1,column=col)
        cd.border = bor
        cd.fill = fillfg
        cd.font  = font_cell
        cd.alignment = aling
        ws.column_dimensions[abc[col-1]].width=20.0

    for col in range(1, 72):
            cd = ws.cell(2,column=col)
            cd.border = bor

    #ws.cell(row=cont,column=1).value = Bd_ANFIBIOSREPTILES.user
    ws.cell(row=cont,column=1).value = Bd_ANFIBIOSREPTILES.codigo_proyecto
    ws.cell(row=cont,column=2).value = Bd_ANFIBIOSREPTILES.fecha_registro
    ws.cell(row=cont,column=3).value = Bd_ANFIBIOSREPTILES.temporada_evaluacion
    ws.cell(row=cont,column=4).value = Bd_ANFIBIOSREPTILES.unidad_vegetacion_quebrada
    ws.cell(row=cont,column=5).value = Bd_ANFIBIOSREPTILES.sector
    ws.cell(row=cont,column=6).value = Bd_ANFIBIOSREPTILES.codigo_quebrada
    ws.cell(row=cont,column=7).value = Bd_ANFIBIOSREPTILES.disciplina_biologica
    ws.cell(row=cont,column=8).value = Bd_ANFIBIOSREPTILES.estacion_muestreo
    ws.cell(row=cont,column=9).value = Bd_ANFIBIOSREPTILES.unidad_muestreo
    ws.cell(row=cont,column=10).value = Bd_ANFIBIOSREPTILES.metodo_muestreo
    ws.cell(row=cont,column=11).value = Bd_ANFIBIOSREPTILES.tamaño_unidad_muestreo
    ws.cell(row=cont,column=12).value = Bd_ANFIBIOSREPTILES.chip_marca_dedos
    ws.cell(row=cont,column=13).value = Bd_ANFIBIOSREPTILES.periodo_dia
    ws.cell(row=cont,column=14).value = Bd_ANFIBIOSREPTILES.hora_inicio_evaluacion
    ws.cell(row=cont,column=15).value = Bd_ANFIBIOSREPTILES.hora_finalizacion_evaluacion
    ws.cell(row=cont,column=16).value = Bd_ANFIBIOSREPTILES.hora_registro
    ws.cell(row=cont,column=17).value = Bd_ANFIBIOSREPTILES.familia
    ws.cell(row=cont,column=18).value = Bd_ANFIBIOSREPTILES.especie
    ws.cell(row=cont,column=19).value = Bd_ANFIBIOSREPTILES.nombre_comun
    ws.cell(row=cont,column=20).value = Bd_ANFIBIOSREPTILES.nombre_local
    ws.cell(row=cont,column=21).value = Bd_ANFIBIOSREPTILES.fuente_clasificacion_taxonomica
    ws.cell(row=cont,column=22).value = Bd_ANFIBIOSREPTILES.x
    ws.cell(row=cont,column=23).value = Bd_ANFIBIOSREPTILES.y
    ws.cell(row=cont,column=24).value = Bd_ANFIBIOSREPTILES.proyeccion
    ws.cell(row=cont,column=25).value = Bd_ANFIBIOSREPTILES.altitud_msnm
    ws.cell(row=cont,column=26).value = Bd_ANFIBIOSREPTILES.datum
    ws.cell(row=cont,column=27).value = Bd_ANFIBIOSREPTILES.temperatura_ambiente
    ws.cell(row=cont,column=28).value = Bd_ANFIBIOSREPTILES.temperatura_suelo
    ws.cell(row=cont,column=29).value = Bd_ANFIBIOSREPTILES.profundidad_hojarrasca_cm
    ws.cell(row=cont,column=30).value = Bd_ANFIBIOSREPTILES.temperie
    ws.cell(row=cont,column=31).value = Bd_ANFIBIOSREPTILES.distancia_perpendicular_unidad_muestreo
    ws.cell(row=cont,column=32).value = Bd_ANFIBIOSREPTILES.altura_registro_distancia_vertical_suelo
    ws.cell(row=cont,column=33).value = Bd_ANFIBIOSREPTILES.tipo_evaluacion
    ws.cell(row=cont,column=34).value = Bd_ANFIBIOSREPTILES.tipo_registro
    ws.cell(row=cont,column=35).value = Bd_ANFIBIOSREPTILES.numero_registro_directo
    ws.cell(row=cont,column=36).value = Bd_ANFIBIOSREPTILES.numero_registro_indirecto
    ws.cell(row=cont,column=37).value = Bd_ANFIBIOSREPTILES.numero_registro_captura
    ws.cell(row=cont,column=38).value = Bd_ANFIBIOSREPTILES.estado_hallazgo
    ws.cell(row=cont,column=39).value = Bd_ANFIBIOSREPTILES.etologia_comportamiento_individuo_registrado
    ws.cell(row=cont,column=40).value = Bd_ANFIBIOSREPTILES.numero_total_individuos
    ws.cell(row=cont,column=41).value = Bd_ANFIBIOSREPTILES.peso
    ws.cell(row=cont,column=42).value = Bd_ANFIBIOSREPTILES.sexo
    ws.cell(row=cont,column=43).value = Bd_ANFIBIOSREPTILES.longitud
    ws.cell(row=cont,column=44).value = Bd_ANFIBIOSREPTILES.longitud_ano_hocico
    ws.cell(row=cont,column=45).value = Bd_ANFIBIOSREPTILES.longitud_cola_lagartijas_serpientes
    ws.cell(row=cont,column=46).value = Bd_ANFIBIOSREPTILES.tipo_reproduccion_anfibios
    ws.cell(row=cont,column=47).value = Bd_ANFIBIOSREPTILES.conducta_reproductva_in_situ
    ws.cell(row=cont,column=48).value = Bd_ANFIBIOSREPTILES.conducta_social
    ws.cell(row=cont,column=49).value = Bd_ANFIBIOSREPTILES.territorialidad
    ws.cell(row=cont,column=50).value = Bd_ANFIBIOSREPTILES.habito
    ws.cell(row=cont,column=51).value = Bd_ANFIBIOSREPTILES.tipo_habitat_in_situ
    ws.cell(row=cont,column=52).value = Bd_ANFIBIOSREPTILES.micro_habitat_in_situ
    ws.cell(row=cont,column=53).value = Bd_ANFIBIOSREPTILES.planta_dominante_microhabitat_in_situ
    ws.cell(row=cont,column=54).value = Bd_ANFIBIOSREPTILES.grupo_trofico
    ws.cell(row=cont,column=55).value = Bd_ANFIBIOSREPTILES.registro_dieta_in_situ
    ws.cell(row=cont,column=56).value = Bd_ANFIBIOSREPTILES.especie_endemica
    ws.cell(row=cont,column=57).value = Bd_ANFIBIOSREPTILES.nivel_endemismo
    ws.cell(row=cont,column=58).value = Bd_ANFIBIOSREPTILES.distribucion_especie_endemica
    ws.cell(row=cont,column=59).value = Bd_ANFIBIOSREPTILES.DSO342004AG
    ws.cell(row=cont,column=60).value = Bd_ANFIBIOSREPTILES.IUCN
    ws.cell(row=cont,column=61).value = Bd_ANFIBIOSREPTILES.CITES
    ws.cell(row=cont,column=62).value = Bd_ANFIBIOSREPTILES.uso_potencial
    ws.cell(row=cont,column=63).value = Bd_ANFIBIOSREPTILES.observaciones_registro
    ws.cell(row=cont,column=64).value = Bd_ANFIBIOSREPTILES.actividad_humana
    ws.cell(row=cont,column=65).value = Bd_ANFIBIOSREPTILES.estado_conservacion_habitat
    ws.cell(row=cont,column=66).value = Bd_ANFIBIOSREPTILES.distancia_componente_minero
    ws.cell(row=cont,column=67).value = Bd_ANFIBIOSREPTILES.medida_control_impedir_ingreso
    ws.cell(row=cont,column=68).value = Bd_ANFIBIOSREPTILES.micro_habitat
    ws.cell(row=cont,column=69).value = Bd_ANFIBIOSREPTILES.animal_instalaciones
    ws.cell(row=cont,column=70).value = Bd_ANFIBIOSREPTILES.riesgo_contacto_animal_instalaciones
    ws.cell(row=cont,column=71).value = Bd_ANFIBIOSREPTILES.observaciones_animales_muertos

    nombre_archivo ="Anfibios_reptiles_BD_2018.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

##----------------------EXPORT_Coordenadas_BD_Resptiles-------------------------#

def Reporte_prueba_coordenadas(request):

    coordenadas = Coordenadas_BD.objects.all()

    wb = Workbook()
    ws = wb.active

    ws['J1'] = 'Georreferenciación WGS 84-18L'

    ws.merge_cells('J1:P1')
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:C2')
    ws.merge_cells('D1:D2')
    ws.merge_cells('E1:E2')
    ws.merge_cells('F1:F2')
    ws.merge_cells('G1:G2')
    ws.merge_cells('H1:H2')
    ws.merge_cells('I1:I2')
    ws.merge_cells('Q1:Q2')
    ws.merge_cells('R1:R2')
    ws.merge_cells('S1:S2')

    #ws['A1'] = 'Usuario'
    ws['A1'] = 'Unidad de Vegetación (Quebrada)'
    ws['B1'] = 'Sector (Alto, medio, bajo)'
    ws['C1'] = 'Codigo de Estación'
    ws['D1'] = 'Estacion de Muestreo '
    ws['E1'] = 'Unidad de Muestreo'
    ws['F1'] = 'Tipo de Unidad Muestral'
    ws['G1'] = 'Tamaño de Unidad Muestral'
    ws['H1'] = 'Subunidad de Muestreo'
    ws['I1'] = 'Disciplina'
    ws['J2'] = 'Inicial-Final UM / Puntual '
    ws['K2'] = 'Este'
    ws['L2'] = 'Norte'
    ws['M2'] = 'Altitud'
    ws['N2'] = 'Este'
    ws['O2'] = 'Norte'
    ws['P2'] = 'Altitud'
    ws['Q1'] = 'Hora (Inicio-Final)'
    ws['R1'] = 'Fecha'
    ws['S1'] = 'Descripción de la estación  Muestreo'
    cont=3

    for coordenada in coordenadas:
        ws.cell(row=cont,column=1).value = coordenada.unidad_vegetacion_quebrada
        ws.cell(row=cont,column=2).value = coordenada.sector
        ws.cell(row=cont,column=3).value = coordenada.codigo_estacion
        ws.cell(row=cont,column=4).value = coordenada.estacion_muestreo
        ws.cell(row=cont,column=5).value = coordenada.unidad_muestreo
        ws.cell(row=cont,column=6).value = coordenada.tipo_unidad_muestral
        ws.cell(row=cont,column=7).value = coordenada.tamaño_unidad_muestreal
        ws.cell(row=cont,column=8).value = coordenada.subunidad_muestreo
        ws.cell(row=cont,column=9).value = coordenada.disciplina
        ws.cell(row=cont,column=10).value = coordenada.georreferenciacion_WGS8418L.inicial_final_puntual
        ws.cell(row=cont,column=11).value = coordenada.georreferenciacion_WGS8418L.este1
        ws.cell(row=cont,column=12).value = coordenada.georreferenciacion_WGS8418L.norte1
        ws.cell(row=cont,column=13).value = coordenada.georreferenciacion_WGS8418L.altitud1
        ws.cell(row=cont,column=14).value = coordenada.georreferenciacion_WGS8418L.este2
        ws.cell(row=cont,column=15).value = coordenada.georreferenciacion_WGS8418L.norte2
        ws.cell(row=cont,column=16).value = coordenada.georreferenciacion_WGS8418L.altitud2
        ws.cell(row=cont,column=17).value = coordenada.hora_inicio_final
        ws.cell(row=cont,column=18,value= coordenada.fecha).number_format = numbers.FORMAT_DATE_XLSX15
        ws.cell(row=cont,column=19).value = coordenada.descripcion_estacion_muestreo
        cont = cont + 1

    nombre_archivo ="Coordenadas_BD_Reptiles_Amfibios.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

#Reporte_Prueba_Coordenada_unico
def Reporte_prueba_coordenadas_unico(request,id):

    coordenada = Coordenadas_BD.objects.get(id=id)


    wb = Workbook()
    ws = wb.active

    ws['J1'] = 'Georreferenciación WGS 84-18L'

    ws.merge_cells('J1:P1')
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:C2')
    ws.merge_cells('D1:D2')
    ws.merge_cells('E1:E2')
    ws.merge_cells('F1:F2')
    ws.merge_cells('G1:G2')
    ws.merge_cells('H1:H2')
    ws.merge_cells('I1:I2')
    ws.merge_cells('Q1:Q2')
    ws.merge_cells('R1:R2')
    ws.merge_cells('S1:S2')

    #ws['A1'] = 'Usuario'
    ws['A1'] = 'Unidad de Vegetación (Quebrada)'
    ws['B1'] = 'Sector (Alto, medio, bajo)'
    ws['C1'] = 'Codigo de Estación'
    ws['D1'] = 'Estacion de Muestreo '
    ws['E1'] = 'Unidad de Muestreo'
    ws['F1'] = 'Tipo de Unidad Muestral'
    ws['G1'] = 'Tamaño de Unidad Muestral'
    ws['H1'] = 'Subunidad de Muestreo'
    ws['I1'] = 'Disciplina'
    ws['J2'] = 'Inicial-Final UM / Puntual '
    ws['K2'] = 'Este'
    ws['L2'] = 'Norte'
    ws['M2'] = 'Altitud'
    ws['N2'] = 'Este'
    ws['O2'] = 'Norte'
    ws['P2'] = 'Altitud'
    ws['Q1'] = 'Hora (Inicio-Final)'
    ws['R1'] = 'Fecha'
    ws['S1'] = 'Descripción de la estación  Muestreo'
    cont=3

    ws.cell(row=cont,column=1).value = coordenada.unidad_vegetacion_quebrada
    ws.cell(row=cont,column=2).value = coordenada.sector
    ws.cell(row=cont,column=3).value = coordenada.codigo_estacion
    ws.cell(row=cont,column=4).value = coordenada.estacion_muestreo
    ws.cell(row=cont,column=5).value = coordenada.unidad_muestreo
    ws.cell(row=cont,column=6).value = coordenada.tipo_unidad_muestral
    ws.cell(row=cont,column=7).value = coordenada.tamaño_unidad_muestreal
    ws.cell(row=cont,column=8).value = coordenada.subunidad_muestreo
    ws.cell(row=cont,column=9).value = coordenada.disciplina
    ws.cell(row=cont,column=10).value = coordenada.georreferenciacion_WGS8418L.inicial_final_puntual
    ws.cell(row=cont,column=11).value = coordenada.georreferenciacion_WGS8418L.este1
    ws.cell(row=cont,column=12).value = coordenada.georreferenciacion_WGS8418L.norte1
    ws.cell(row=cont,column=13).value = coordenada.georreferenciacion_WGS8418L.altitud1
    ws.cell(row=cont,column=14).value = coordenada.georreferenciacion_WGS8418L.este2
    ws.cell(row=cont,column=15).value = coordenada.georreferenciacion_WGS8418L.norte2
    ws.cell(row=cont,column=16).value = coordenada.georreferenciacion_WGS8418L.altitud2
    ws.cell(row=cont,column=17).value = coordenada.hora_inicio_final
    ws.cell(row=cont,column=18,value= coordenada.fecha).number_format = numbers.FORMAT_DATE_XLSX15
    ws.cell(row=cont,column=19).value = coordenada.descripcion_estacion_muestreo

    nombre_archivo ="Coordenadas_BD_Reptiles_Amfibios.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response


##----------------------EXPORT_Fotos_BD_Resptiles-------------------------#

def export_BD_Fotos_Reptiles_Anfibios_xls(request):

    fotos = BD_fotos_BD.objects.all()

    wb = Workbook()
    ws = wb.active
    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble
    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    bor = Border(top=double, left=thin, right=thin, bottom=double)
    fillfg = PatternFill("solid", fgColor="00969696")
    fil = fill = GradientFill(stop=("000000", "FFFFFF"))
    font_cell = Font(b=True, color="000000")
    aling = Alignment(horizontal="center", vertical="center")

    #definicion de celdas(cabezera) con sus estilos
    a1 = ws['A1']
    a1.value = 'Lugar de Toma de Foto'
    a1.border = bor
    a1.fill = fillfg
    a1.font  = font_cell
    a1.alignment = aling

    g1 = ws['G1']
    g1.value = 'Codificación'
    g1.border = bor
    g1.fill = fillfg
    g1.font  = font_cell
    g1.alignment = aling

    i1 = ws['I1']
    i1.value = 'Tipo de Foto'
    i1.border = bor
    i1.fill = fillfg
    i1.font  = font_cell
    i1.alignment = aling

    k1 = ws['K1']
    k1.value = 'Especie'
    k1.border = bor
    k1.fill = fillfg
    k1.font  = font_cell
    k1.alignment = aling

    m1 = ws['M1']
    m1.value = 'Ubicación'
    m1.border = bor
    m1.fill = fillfg
    m1.font  = font_cell
    m1.alignment = aling

    q1 = ws['Q1']
    q1.value = 'Breve descripcion de la foto'
    q1.border = bor
    q1.fill = fillfg
    q1.font  = font_cell
    q1.alignment = aling

    #union de celdas
    ws.merge_cells('A1:F1')
    ws.merge_cells('G1:H1')
    ws.merge_cells('I1:J1')
    ws.merge_cells('K1:L1')
    ws.merge_cells('M1:P1')
    ws.merge_cells('Q1:Q2')

    #cabezera principal
    ws['A2'] = 'N°'
    ws['B2'] = 'Unidad de Vegetación '
    ws['C2'] = 'Sector (Alto, medio, bajo)'
    ws['D2'] = 'Codigo de la Quebrada'
    ws['E2'] = 'Estacion de Muestreo '
    ws['F2'] = 'Unidad de Muestreo (UM)'
    ws['G2'] = 'Código de la foto'
    ws['H2'] = 'Fecha de la Toma'
    ws['I2'] = 'Especie'
    ws['J2'] = 'Paisaje '
    ws['K2'] = 'Nombre científico'
    ws['L2'] = 'Nombre común'
    ws['M2'] = 'Localidad'
    ws['N2'] = 'Distrito'
    ws['O2'] = 'Provincia'
    ws['P2'] = 'Departamento'
    ws.column_dimensions["Q"].width=40.0

    abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']

    for col in range(1, 17):
        cd = ws.cell(row=2,column=col)
        cd.border = bor
        cd.fill = fillfg
        cd.font  = font_cell
        cd.alignment = aling
        if(col == 3 or col == 5 or col == 6):
            ws.column_dimensions[abc[col-1]].width=25.0
        else:
            ws.column_dimensions[abc[col-1]].width=20.0

    cont=3

    for foto in fotos:
        ws.cell(row=cont,column=1).value = foto.lugar_toma_foto.n
        ws.cell(row=cont,column=2).value = foto.lugar_toma_foto.unidad_vegetacion
        ws.cell(row=cont,column=3).value = foto.lugar_toma_foto.sector_alto_bajo_medio
        ws.cell(row=cont,column=4).value = foto.lugar_toma_foto.codigo_quebrada
        ws.cell(row=cont,column=5).value = foto.lugar_toma_foto.estacion_muestreo
        ws.cell(row=cont,column=6).value = foto.lugar_toma_foto.unidad_muestreo_um
        ws.cell(row=cont,column=7).value = foto.codificacion.codigo_foto
        ws.cell(row=cont,column=8,value = foto.codificacion.fecha_toma).number_format = numbers.FORMAT_DATE_XLSX15
        ws.cell(row=cont,column=9).value = foto.tipo_foto.especie
        ws.cell(row=cont,column=10).value = foto.tipo_foto.paisaje
        ws.cell(row=cont,column=11).value = foto.especie.nombre_cientifico
        ws.cell(row=cont,column=12).value = foto.especie.nombre_comun
        ws.cell(row=cont,column=13).value = foto.ubicacion.localidad
        ws.cell(row=cont,column=14).value = foto.ubicacion.distrito
        ws.cell(row=cont,column=15).value = foto.ubicacion.provincia
        ws.cell(row=cont,column=16).value = foto.ubicacion.departamento
        ws.cell(row=cont,column=17).value = foto.breve_descripcion_foto
        cont = cont + 1

    nombre_archivo ="Fotos_BD_Resptiles.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

#Fotos_reptiles_unico
def export_fotos_Reptiles_Anfibios_unico(request,id):

    foto = BD_fotos_BD.objects.get(id=id)

    wb = Workbook()
    ws = wb.active
    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble
    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    bor = Border(top=double, left=thin, right=thin, bottom=double)
    fillfg = PatternFill("solid", fgColor="00969696")
    fil = fill = GradientFill(stop=("000000", "FFFFFF"))
    font_cell = Font(b=True, color="000000")
    aling = Alignment(horizontal="center", vertical="center")

    #definicion de celdas(cabezera) con sus estilos
    a1 = ws['A1']
    a1.value = 'Lugar de Toma de Foto'
    a1.border = bor
    a1.fill = fillfg
    a1.font  = font_cell
    a1.alignment = aling

    g1 = ws['G1']
    g1.value = 'Codificación'
    g1.border = bor
    g1.fill = fillfg
    g1.font  = font_cell
    g1.alignment = aling

    i1 = ws['I1']
    i1.value = 'Tipo de Foto'
    i1.border = bor
    i1.fill = fillfg
    i1.font  = font_cell
    i1.alignment = aling

    k1 = ws['K1']
    k1.value = 'Especie'
    k1.border = bor
    k1.fill = fillfg
    k1.font  = font_cell
    k1.alignment = aling

    m1 = ws['M1']
    m1.value = 'Ubicación'
    m1.border = bor
    m1.fill = fillfg
    m1.font  = font_cell
    m1.alignment = aling

    q1 = ws['Q1']
    q1.value = 'Breve descripcion de la foto'
    q1.border = bor
    q1.fill = fillfg
    q1.font  = font_cell
    q1.alignment = aling

    #union de celdas
    ws.merge_cells('A1:F1')
    ws.merge_cells('G1:H1')
    ws.merge_cells('I1:J1')
    ws.merge_cells('K1:L1')
    ws.merge_cells('M1:P1')
    ws.merge_cells('Q1:Q2')

    #cabezera principal
    ws['A2'] = 'N°'
    ws['B2'] = 'Unidad de Vegetación '
    ws['C2'] = 'Sector (Alto, medio, bajo)'
    ws['D2'] = 'Codigo de la Quebrada'
    ws['E2'] = 'Estacion de Muestreo '
    ws['F2'] = 'Unidad de Muestreo (UM)'
    ws['G2'] = 'Código de la foto'
    ws['H2'] = 'Fecha de la Toma'
    ws['I2'] = 'Especie'
    ws['J2'] = 'Paisaje '
    ws['K2'] = 'Nombre científico'
    ws['L2'] = 'Nombre común'
    ws['M2'] = 'Localidad'
    ws['N2'] = 'Distrito'
    ws['O2'] = 'Provincia'
    ws['P2'] = 'Departamento'
    ws.column_dimensions["Q"].width=40.0

    abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z']

    for col in range(1, 17):
        cd = ws.cell(row=2,column=col)
        cd.border = bor
        cd.fill = fillfg
        cd.font  = font_cell
        cd.alignment = aling
        if(col == 3 or col == 5 or col == 6):
            ws.column_dimensions[abc[col-1]].width=25.0
        else:
            ws.column_dimensions[abc[col-1]].width=20.0

    cont=3


    ws.cell(row=cont,column=1).value = foto.lugar_toma_foto.n
    ws.cell(row=cont,column=2).value = foto.lugar_toma_foto.unidad_vegetacion
    ws.cell(row=cont,column=3).value = foto.lugar_toma_foto.sector_alto_bajo_medio
    ws.cell(row=cont,column=4).value = foto.lugar_toma_foto.codigo_quebrada
    ws.cell(row=cont,column=5).value = foto.lugar_toma_foto.estacion_muestreo
    ws.cell(row=cont,column=6).value = foto.lugar_toma_foto.unidad_muestreo_um
    ws.cell(row=cont,column=7).value = foto.codificacion.codigo_foto
    ws.cell(row=cont,column=8,value = foto.codificacion.fecha_toma).number_format = numbers.FORMAT_DATE_XLSX15
    ws.cell(row=cont,column=9).value = foto.tipo_foto.especie
    ws.cell(row=cont,column=10).value = foto.tipo_foto.paisaje
    ws.cell(row=cont,column=11).value = foto.especie.nombre_cientifico
    ws.cell(row=cont,column=12).value = foto.especie.nombre_comun
    ws.cell(row=cont,column=13).value = foto.ubicacion.localidad
    ws.cell(row=cont,column=14).value = foto.ubicacion.distrito
    ws.cell(row=cont,column=15).value = foto.ubicacion.provincia
    ws.cell(row=cont,column=16).value = foto.ubicacion.departamento
    ws.cell(row=cont,column=17).value = foto.breve_descripcion_foto

    nombre_archivo ="Fotos_BD_Resptiles.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response


##----------------------EXPORT_Hoja1_BD_Resptiles-------------------------#

def export_BD_Hoja1_Reptiles_Anfibios_xls(request):

    hojas1 = hoja1_BD.objects.all()

    wb = Workbook()
    ws = wb.active

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble
    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    bor = Border(top=double, left=thin, right=thin, bottom=double)
    fillfg = PatternFill("solid", fgColor="00969696")
    fil = fill = GradientFill(stop=("000000", "FFFFFF"))
    font_cell = Font(b=True, color="000000")
    aling = Alignment(horizontal="center", vertical="center")

    a1 = ws['A1']
    a1.value = 'Unidad de vegetación'
    a1.border = bor
    a1.fill = fillfg
    a1.font  = font_cell
    a1.alignment = aling

    b1 = ws['B1']
    b1.value = 'Estación de monitoreo'
    b1.border = bor
    b1.fill = fillfg
    b1.font  = font_cell
    b1.alignment = aling

    c1 = ws['C1']
    c1.value = 'Coordenadas UTM (WGS84)'
    c1.border = bor
    c1.fill = fillfg
    c1.font  = font_cell
    c1.alignment = aling

    c2 = ws['C2']
    c2.value = 'Inicio'
    c2.border = bor
    c2.fill = fillfg
    c2.font  = font_cell
    c2.alignment = aling

    e2 = ws['E2']
    e2.value = 'Final'
    e2.border = bor
    e2.fill = fillfg
    e2.font  = font_cell
    e2.alignment = aling

    c3 = ws['C3']
    c3.value = 'Este'
    c3.border = bor
    c3.fill = fillfg
    c3.font  = font_cell
    c3.alignment = aling

    d3 = ws['D3']
    d3.value = 'Norte'
    d3.border = bor
    d3.fill = fillfg
    d3.font  = font_cell
    d3.alignment = aling

    e3 = ws['E3']
    e3.value = 'Este'
    e3.border = bor
    e3.fill = fillfg
    e3.font  = font_cell
    e3.alignment = aling

    f3 = ws['F3']
    f3.value = 'Norte'
    f3.border = bor
    f3.fill = fillfg
    f3.font  = font_cell
    f3.alignment = aling

    ws.merge_cells('A1:A3')
    ws.merge_cells('B1:B3')
    ws.merge_cells('C1:F1')
    ws.merge_cells('C2:D2')
    ws.merge_cells('E2:F2')

    ws.column_dimensions["A"].width=20.0
    ws.column_dimensions["B"].width=20.0

    cont=4

    for hoja in hojas1:
        ws.cell(row=cont,column=1).value = hoja.unidad_vegetacion
        ws.cell(row=cont,column=2).value = hoja.estacion_monitoreo
        ws.cell(row=cont,column=3).value = hoja.coordenadas_inicio.este
        ws.cell(row=cont,column=4).value = hoja.coordenadas_inicio.norte
        ws.cell(row=cont,column=5).value = hoja.coordenadas_final.este
        ws.cell(row=cont,column=6).value = hoja.coordenadas_final.norte
        cont = cont + 1

    nombre_archivo ="Hoja1_BD_Resptiles.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

#Hoja1_reptiles Unico
def export_BD_Hoja1_Reptiles_Anfibios_xls_unic(request,id):

    hoja = hoja1_BD.objects.get(id=id)

    wb = Workbook()
    ws = wb.active

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble
    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    bor = Border(top=double, left=thin, right=thin, bottom=double)
    fillfg = PatternFill("solid", fgColor="00969696")
    fil = fill = GradientFill(stop=("000000", "FFFFFF"))
    font_cell = Font(b=True, color="000000")
    aling = Alignment(horizontal="center", vertical="center")

    a1 = ws['A1']
    a1.value = 'Unidad de vegetación'
    a1.border = bor
    a1.fill = fillfg
    a1.font  = font_cell
    a1.alignment = aling

    b1 = ws['B1']
    b1.value = 'Estación de monitoreo'
    b1.border = bor
    b1.fill = fillfg
    b1.font  = font_cell
    b1.alignment = aling

    c1 = ws['C1']
    c1.value = 'Coordenadas UTM (WGS84)'
    c1.border = bor
    c1.fill = fillfg
    c1.font  = font_cell
    c1.alignment = aling

    c2 = ws['C2']
    c2.value = 'Inicio'
    c2.border = bor
    c2.fill = fillfg
    c2.font  = font_cell
    c2.alignment = aling

    e2 = ws['E2']
    e2.value = 'Final'
    e2.border = bor
    e2.fill = fillfg
    e2.font  = font_cell
    e2.alignment = aling

    c3 = ws['C3']
    c3.value = 'Este'
    c3.border = bor
    c3.fill = fillfg
    c3.font  = font_cell
    c3.alignment = aling

    d3 = ws['D3']
    d3.value = 'Norte'
    d3.border = bor
    d3.fill = fillfg
    d3.font  = font_cell
    d3.alignment = aling

    e3 = ws['E3']
    e3.value = 'Este'
    e3.border = bor
    e3.fill = fillfg
    e3.font  = font_cell
    e3.alignment = aling

    f3 = ws['F3']
    f3.value = 'Norte'
    f3.border = bor
    f3.fill = fillfg
    f3.font  = font_cell
    f3.alignment = aling

    ws.merge_cells('A1:A3')
    ws.merge_cells('B1:B3')
    ws.merge_cells('C1:F1')
    ws.merge_cells('C2:D2')
    ws.merge_cells('E2:F2')

    ws.column_dimensions["A"].width=20.0
    ws.column_dimensions["B"].width=20.0

    cont=4

    ws.cell(row=cont,column=1).value = hoja.unidad_vegetacion
    ws.cell(row=cont,column=2).value = hoja.estacion_monitoreo
    ws.cell(row=cont,column=3).value = hoja.coordenadas_inicio.este
    ws.cell(row=cont,column=4).value = hoja.coordenadas_inicio.norte
    ws.cell(row=cont,column=5).value = hoja.coordenadas_final.este
    ws.cell(row=cont,column=6).value = hoja.coordenadas_final.norte


    nombre_archivo ="Hoja1_BD_Resptiles_unico.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

# ============================== SECCION MAMIFEROS =====================================

#Base_Datos_Mamiferos
def export_base_datos_mamiferos(request):

    Datos = BD_mamiferos_BD_Mamiferos.objects.all()

    wb = Workbook()
    ws = wb.active

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble
    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    bor = Border(top=double, left=thin, right=thin, bottom=double)
    fillfg = PatternFill("solid", fgColor="00969696")
    fil = fill = GradientFill(stop=("000000", "FFFFFF"))
    font_cell = Font(b=True, color="000000")
    aling = Alignment(horizontal="center", vertical="center")


    #ws['A1'] = 'Usuario'
    ws['A1'] = 'Codigo de proyecto'
    ws['B1'] = 'Fecha de registro'
    ws['C1'] = 'Temporada'
    ws['D1'] = 'Unidad de vegetación'
    ws['E1'] = 'Disciplina biologia'
    ws['F1'] = 'Estacion de muestreo'
    ws['G1'] = 'Unidad de muestreo'
    ws['H1'] = 'Subunidad de muestreo'
    ws['I1'] = 'Metodo de muestreo'
    ws['J1'] = 'Tamaño de la unidad de muestreo(m)'
    ws['K1'] = 'N° Estaciones de muestreo (trampas, redes, etc.)'
    ws['L1'] = 'Periodo del dia'
    ws['M1'] = 'Hora de inicio de evaluación'
    ws['N1'] = 'Hora de finalización de evaluación'
    ws['O1'] = 'Hora de registro'
    ws['P1'] = 'Orden'
    ws['Q1'] = 'Familia'
    ws['R1'] = 'Especie'
    ws['S1'] = 'Nombre común'
    ws['T1'] = 'Nombre local'
    ws['U1'] = 'Fuente de clasificación taxonomica'
    ws['V1'] = 'X'
    ws['W1'] = 'Y'
    ws['X1'] = 'Proyección'
    ws['Y1'] = 'Altitud (msnm)'
    ws['Z1'] = 'Datum'
    ws['AA1'] = 'Distancia de la unidad de muestreo'
    ws['AB1'] = 'Distancia perpendicular de la unidad de muestreo'
    ws['AC1'] = 'Tipo de Evaluación'
    ws['AD1'] = 'Tipo de registro'
    ws['AE1'] = 'N° Registro Directo'
    ws['AF1'] = 'N°Registro indirecto'
    ws['AG1'] = 'N°Registro por captura (murcis, roedores)'
    ws['AH1'] = 'Estado del hallazgo'
    ws['AI1'] = 'Observaciones (etología)'
    ws['AJ1'] = 'Nº total de individuos'
    ws['AK1'] = 'Crias'
    ws['AL1'] = 'Infantes'
    ws['AM1'] = 'Juvenil'
    ws['AN1'] = 'Adulto'
    ws['AO1'] = 'Peso (gr)'
    ws['AP1'] = 'Sexo'
    ws['AQ1'] = 'Longitud total (cm)'
    ws['AR1'] = 'Longitud antebrazo (cm)'
    ws['AS1'] = 'Longitud tibia (cm)'
    ws['AT1'] = 'Longitud pata (cm)'
    ws['AU1'] = 'Longitud oreja (cm)'
    ws['AV1'] = 'Longitud tragus (cm)'
    ws['AW1'] = 'Longitud cola (cm)'
    ws['AX1'] = 'Estado reproductivo'
    ws['AY1'] = 'Conducta reproductiva'
    ws['AZ1'] = 'Conducta social'
    ws['BA1'] = 'Especie según Ambiente'
    ws['BB1'] = 'Tipo de locomoción'
    ws['BC1'] = 'Tipo de Hábitat'
    ws['BD1'] = 'Micro-habitat'
    ws['BE1'] = 'Grupo Trófico'
    ws['BF1'] = 'Registro de dieta '
    ws['BG1'] = 'Especie Endémica'
    ws['BH1'] = 'Nivel de Endemismo'
    ws['BI1'] = 'Zona de Endemismo'
    ws['BJ1'] = 'DS O04-2014-AG'
    ws['BK1'] = 'IUCN'
    ws['BL1'] = 'CITES'
    ws['BM1'] = 'Uso potencial'
    ws['BN1'] = 'Observaciones sobre la muestra'
    ws['BO1'] = 'Temperie'
    ws['BP1'] = 'Actividad humana'
    ws['BQ1'] = 'Estado de conservación del hábitat'
    ws['BR1'] = 'Colector de registro'
    ws['BS1'] = 'Codigo de Colecta'
    ws['BT1'] = 'Tipo de muestra'
    ws['BU1'] = 'Estado de la muestra'
    ws['BV1'] = 'Tipo de preparación de la muestra'
    ws['BW1'] = 'Institución de depósito'
    ws['BX1'] = 'Código fotos'
    ws['BY1'] = 'Obs'

    cont=3

    abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
               'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ',
               'BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR','BS','BT','BU','BV','BW','BX','BY','BZ']

    for col in range (1,77):
        cd = ws.cell(row=1,column=col)
        cd.border = bor
        cd.fill = fillfg
        cd.font  = font_cell
        cd.alignment = aling
        ws.column_dimensions[abc[col-1]].width=20.0

    for col in range(1, 77):
            for row in range(1,len(Datos)+1):
                cd = ws.cell(row=row+1,column=col)
                cd.border = bor

    for bd_mamifero in Datos:
        ws.cell(row=cont,column=1).value = bd_mamifero.codigo_proyecto
        ws.cell(row=cont,column=2,value = bd_mamifero.fecha_registro).number_format = numbers.FORMAT_DATE_XLSX15
        ws.cell(row=cont,column=3).value = bd_mamifero.temporada
        ws.cell(row=cont,column=4).value = bd_mamifero.unidad_vegetacion
        ws.cell(row=cont,column=5).value = bd_mamifero.disciplina_biologica
        ws.cell(row=cont,column=6).value = bd_mamifero.estacion_muestreo
        ws.cell(row=cont,column=7).value = bd_mamifero.unidad_muestreo
        ws.cell(row=cont,column=8).value = bd_mamifero.subunidad_muestreo
        ws.cell(row=cont,column=9).value = bd_mamifero.metodo_muestreo
        ws.cell(row=cont,column=10).value = bd_mamifero.tamaño_unidad_muestreo
        ws.cell(row=cont,column=11).value = bd_mamifero.numero_estaciones_muestreo
        ws.cell(row=cont,column=12).value = bd_mamifero.periodo_dia
        ws.cell(row=cont,column=13).value = bd_mamifero.hora_inicio_evaluacion
        ws.cell(row=cont,column=14).value = bd_mamifero.hora_finalizacion_evaluacion
        ws.cell(row=cont,column=15).value = bd_mamifero.hora_registro
        ws.cell(row=cont,column=16).value = bd_mamifero.orden
        ws.cell(row=cont,column=17).value = bd_mamifero.familia
        ws.cell(row=cont,column=18).value = bd_mamifero.especie
        ws.cell(row=cont,column=19).value = bd_mamifero.nombre_comun
        ws.cell(row=cont,column=20).value = bd_mamifero.nombre_local
        ws.cell(row=cont,column=21).value = bd_mamifero.fuente_clasificacion_taxonomica
        ws.cell(row=cont,column=22).value = bd_mamifero.x
        ws.cell(row=cont,column=23).value = bd_mamifero.y
        ws.cell(row=cont,column=24).value = bd_mamifero.proyeccion
        ws.cell(row=cont,column=25).value = bd_mamifero.altitud_msnm
        ws.cell(row=cont,column=26).value = bd_mamifero.datum
        ws.cell(row=cont,column=27).value = bd_mamifero.distancia_unidad_muestreo
        ws.cell(row=cont,column=28).value = bd_mamifero.distancia_perpendicular_unidad_muestreo
        ws.cell(row=cont,column=29).value = bd_mamifero.tipo_evaluacion
        ws.cell(row=cont,column=30).value = bd_mamifero.tipo_registro
        ws.cell(row=cont,column=31).value = bd_mamifero.numero_registro_directo
        ws.cell(row=cont,column=32).value = bd_mamifero.numero_registro_indirecto
        ws.cell(row=cont,column=33).value = bd_mamifero.registro_captura
        ws.cell(row=cont,column=34).value = bd_mamifero.estado_hallazgo
        ws.cell(row=cont,column=35).value = bd_mamifero.observaciones_etologia
        ws.cell(row=cont,column=36).value = bd_mamifero.numero_total_individuos
        ws.cell(row=cont,column=37).value = bd_mamifero.crias
        ws.cell(row=cont,column=38).value = bd_mamifero.infantes
        ws.cell(row=cont,column=39).value = bd_mamifero.juvenil
        ws.cell(row=cont,column=40).value = bd_mamifero.adulto
        ws.cell(row=cont,column=41).value = bd_mamifero.peso
        ws.cell(row=cont,column=42).value = bd_mamifero.sexo
        ws.cell(row=cont,column=43).value = bd_mamifero.longitud_total
        ws.cell(row=cont,column=44).value = bd_mamifero.longitud_antebrazo
        ws.cell(row=cont,column=45).value = bd_mamifero.longitud_tibia
        ws.cell(row=cont,column=46).value = bd_mamifero.longitud_pata
        ws.cell(row=cont,column=47).value = bd_mamifero.longitud_oreja
        ws.cell(row=cont,column=48).value = bd_mamifero.longitud_tragus
        ws.cell(row=cont,column=49).value = bd_mamifero.longitud_cola
        ws.cell(row=cont,column=50).value = bd_mamifero.estado_reproductivo
        ws.cell(row=cont,column=51).value = bd_mamifero.conducta_reproductiva
        ws.cell(row=cont,column=52).value = bd_mamifero.conducta_social
        ws.cell(row=cont,column=53).value = bd_mamifero.especie_segun_ambiente
        ws.cell(row=cont,column=54).value = bd_mamifero.tipo_locomocion
        ws.cell(row=cont,column=55).value = bd_mamifero.tipo_habitat
        ws.cell(row=cont,column=56).value = bd_mamifero.micro_habitat
        ws.cell(row=cont,column=57).value = bd_mamifero.grupo_trofico
        ws.cell(row=cont,column=58).value = bd_mamifero.registro_dieta
        ws.cell(row=cont,column=59).value = bd_mamifero.especie_endemica
        ws.cell(row=cont,column=60).value = bd_mamifero.nivel_endemismo
        ws.cell(row=cont,column=61).value = bd_mamifero.zona_endemismo
        ws.cell(row=cont,column=62).value = bd_mamifero.DSO042014AG
        ws.cell(row=cont,column=63).value = bd_mamifero.IUCN
        ws.cell(row=cont,column=64).value = bd_mamifero.CITES
        ws.cell(row=cont,column=65).value = bd_mamifero.uso_potencial
        ws.cell(row=cont,column=66).value = bd_mamifero.observaciones_sobre_muestra
        ws.cell(row=cont,column=67).value = bd_mamifero.temperie
        ws.cell(row=cont,column=68).value = bd_mamifero.actividad_humana
        ws.cell(row=cont,column=69).value = bd_mamifero.estado_conservacion_habitat
        ws.cell(row=cont,column=70).value = bd_mamifero.colector_registro
        ws.cell(row=cont,column=71).value = bd_mamifero.codigo_colecta
        ws.cell(row=cont,column=72).value = bd_mamifero.tipo_muestra
        ws.cell(row=cont,column=73).value = bd_mamifero.estado_muestra
        ws.cell(row=cont,column=74).value = bd_mamifero.tipo_preparacion_muestra
        ws.cell(row=cont,column=75).value = bd_mamifero.institucion_deposito
        ws.cell(row=cont,column=76).value = bd_mamifero.codigo_fotos
        ws.cell(row=cont,column=77).value = bd_mamifero.observaciones
        cont = cont + 1

    nombre_archivo ="BD_Mamiferos.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

#BD_Mamiferos Unico
def export_base_datos_mamiferos_unic(request,id):

    bd_mamifero = BD_mamiferos_BD_Mamiferos.objects.get(id=id)

    wb = Workbook()
    ws = wb.active

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble
    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    bor = Border(top=double, left=thin, right=thin, bottom=double)
    fillfg = PatternFill("solid", fgColor="00969696")
    fil = fill = GradientFill(stop=("000000", "FFFFFF"))
    font_cell = Font(b=True, color="000000")
    aling = Alignment(horizontal="center", vertical="center")

    ws['A1'] = 'Codigo de proyecto'
    ws['B1'] = 'Fecha de registro'
    ws['C1'] = 'Temporada'
    ws['D1'] = 'Unidad de vegetación'
    ws['E1'] = 'Disciplina biologia'
    ws['F1'] = 'Estacion de muestreo'
    ws['G1'] = 'Unidad de muestreo'
    ws['H1'] = 'Subunidad de muestreo'
    ws['I1'] = 'Metodo de muestreo'
    ws['J1'] = 'Tamaño de la unidad de muestreo(m)'
    ws['K1'] = 'N° Estaciones de muestreo (trampas, redes, etc.)'
    ws['L1'] = 'Periodo del dia'
    ws['M1'] = 'Hora de inicio de evaluación'
    ws['N1'] = 'Hora de finalización de evaluación'
    ws['O1'] = 'Hora de registro'
    ws['P1'] = 'Orden'
    ws['Q1'] = 'Familia'
    ws['R1'] = 'Especie'
    ws['S1'] = 'Nombre común'
    ws['T1'] = 'Nombre local'
    ws['U1'] = 'Fuente de clasificación taxonomica'
    ws['V1'] = 'X'
    ws['W1'] = 'Y'
    ws['X1'] = 'Proyección'
    ws['Y1'] = 'Altitud (msnm)'
    ws['Z1'] = 'Datum'

    ws['AA1'] = 'Distancia de la unidad de muestreo'
    ws['AB1'] = 'Distancia perpendicular de la unidad de muestreo'
    ws['AC1'] = 'Tipo de Evaluación'
    ws['AD1'] = 'Tipo de registro'
    ws['AE1'] = 'N° Registro Directo'
    ws['AF1'] = 'N°Registro indirecto'
    ws['AG1'] = 'N°Registro por captura (murcis, roedores)'
    ws['AH1'] = 'Estado del hallazgo'
    ws['AI1'] = 'Observaciones (etología)'
    ws['AJ1'] = 'Nº total de individuos'
    ws['AK1'] = 'Crias'
    ws['AL1'] = 'Infantes'
    ws['AM1'] = 'Juvenil'
    ws['AN1'] = 'Adulto'
    ws['AO1'] = 'Peso (gr)'
    ws['AP1'] = 'Sexo'
    ws['AQ1'] = 'Longitud total (cm)'
    ws['AR1'] = 'Longitud antebrazo (cm)'
    ws['AS1'] = 'Longitud tibia (cm)'
    ws['AT1'] = 'Longitud pata (cm)'
    ws['AU1'] = 'Longitud oreja (cm)'
    ws['AV1'] = 'Longitud tragus (cm)'
    ws['AW1'] = 'Longitud cola (cm)'
    ws['AX1'] = 'Estado reproductivo'
    ws['AY1'] = 'Conducta reproductiva'
    ws['AZ1'] = 'Conducta social'

    ws['BA1'] = 'Especie según Ambiente'
    ws['BB1'] = 'Tipo de locomoción'
    ws['BC1'] = 'Tipo de Hábitat'
    ws['BD1'] = 'Micro-habitat'
    ws['BE1'] = 'Grupo Trófico'
    ws['BF1'] = 'Registro de dieta '
    ws['BG1'] = 'Especie Endémica'
    ws['BH1'] = 'Nivel de Endemismo'
    ws['BI1'] = 'Zona de Endemismo'
    ws['BJ1'] = 'DS O04-2014-AG'
    ws['BK1'] = 'IUCN'
    ws['BL1'] = 'CITES'
    ws['BM1'] = 'Uso potencial'
    ws['BN1'] = 'Observaciones sobre la muestra'
    ws['BO1'] = 'Temperie'
    ws['BP1'] = 'Actividad humana'
    ws['BQ1'] = 'Estado de conservación del hábitat'
    ws['BR1'] = 'Colector de registro'
    ws['BS1'] = 'Codigo de Colecta'
    ws['BT1'] = 'Tipo de muestra'
    ws['BU1'] = 'Estado de la muestra'
    ws['BV1'] = 'Tipo de preparación de la muestra'
    ws['BW1'] = 'Institución de depósito'

    ws['BX1'] = 'Código fotos'
    ws['BY1'] = 'Obs'

    cont=2

    abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
                   'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ',
                   'BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR','BS','BT','BU','BV','BW','BX','BY','BZ']

    for col in range (1,77):
        cd = ws.cell(row=1,column=col)
        cd.border = bor
        cd.fill = fillfg
        cd.font  = font_cell
        cd.alignment = aling
        ws.column_dimensions[abc[col-1]].width=20.0

    for col in range(1, 77):
            cd = ws.cell(2,column=col)
            cd.border = bor

    ws.cell(row=cont,column=1).value = bd_mamifero.codigo_proyecto
    ws.cell(row=cont,column=2,value = bd_mamifero.fecha_registro).number_format = numbers.FORMAT_DATE_XLSX15
    ws.cell(row=cont,column=3).value = bd_mamifero.temporada
    ws.cell(row=cont,column=4).value = bd_mamifero.unidad_vegetacion
    ws.cell(row=cont,column=5).value = bd_mamifero.disciplina_biologica
    ws.cell(row=cont,column=6).value = bd_mamifero.estacion_muestreo
    ws.cell(row=cont,column=7).value = bd_mamifero.unidad_muestreo
    ws.cell(row=cont,column=8).value = bd_mamifero.subunidad_muestreo
    ws.cell(row=cont,column=9).value = bd_mamifero.metodo_muestreo
    ws.cell(row=cont,column=10).value = bd_mamifero.tamaño_unidad_muestreo
    ws.cell(row=cont,column=11).value = bd_mamifero.numero_estaciones_muestreo
    ws.cell(row=cont,column=12).value = bd_mamifero.periodo_dia
    ws.cell(row=cont,column=13).value = bd_mamifero.hora_inicio_evaluacion
    ws.cell(row=cont,column=14).value = bd_mamifero.hora_finalizacion_evaluacion
    ws.cell(row=cont,column=15).value = bd_mamifero.hora_registro
    ws.cell(row=cont,column=16).value = bd_mamifero.orden
    ws.cell(row=cont,column=17).value = bd_mamifero.familia
    ws.cell(row=cont,column=18).value = bd_mamifero.especie
    ws.cell(row=cont,column=19).value = bd_mamifero.nombre_comun
    ws.cell(row=cont,column=20).value = bd_mamifero.nombre_local
    ws.cell(row=cont,column=21).value = bd_mamifero.fuente_clasificacion_taxonomica
    ws.cell(row=cont,column=22).value = bd_mamifero.x
    ws.cell(row=cont,column=23).value = bd_mamifero.y
    ws.cell(row=cont,column=24).value = bd_mamifero.proyeccion
    ws.cell(row=cont,column=25).value = bd_mamifero.altitud_msnm
    ws.cell(row=cont,column=26).value = bd_mamifero.datum
    ws.cell(row=cont,column=27).value = bd_mamifero.distancia_unidad_muestreo
    ws.cell(row=cont,column=28).value = bd_mamifero.distancia_perpendicular_unidad_muestreo
    ws.cell(row=cont,column=29).value = bd_mamifero.tipo_evaluacion
    ws.cell(row=cont,column=30).value = bd_mamifero.tipo_registro
    ws.cell(row=cont,column=31).value = bd_mamifero.numero_registro_directo
    ws.cell(row=cont,column=32).value = bd_mamifero.numero_registro_indirecto
    ws.cell(row=cont,column=33).value = bd_mamifero.registro_captura
    ws.cell(row=cont,column=34).value = bd_mamifero.estado_hallazgo
    ws.cell(row=cont,column=35).value = bd_mamifero.observaciones_etologia
    ws.cell(row=cont,column=36).value = bd_mamifero.numero_total_individuos
    ws.cell(row=cont,column=37).value = bd_mamifero.crias
    ws.cell(row=cont,column=38).value = bd_mamifero.infantes
    ws.cell(row=cont,column=39).value = bd_mamifero.juvenil
    ws.cell(row=cont,column=40).value = bd_mamifero.adulto
    ws.cell(row=cont,column=41).value = bd_mamifero.peso
    ws.cell(row=cont,column=42).value = bd_mamifero.sexo
    ws.cell(row=cont,column=43).value = bd_mamifero.longitud_total
    ws.cell(row=cont,column=44).value = bd_mamifero.longitud_antebrazo
    ws.cell(row=cont,column=45).value = bd_mamifero.longitud_tibia
    ws.cell(row=cont,column=46).value = bd_mamifero.longitud_pata
    ws.cell(row=cont,column=47).value = bd_mamifero.longitud_oreja
    ws.cell(row=cont,column=48).value = bd_mamifero.longitud_tragus
    ws.cell(row=cont,column=49).value = bd_mamifero.longitud_cola
    ws.cell(row=cont,column=50).value = bd_mamifero.estado_reproductivo
    ws.cell(row=cont,column=51).value = bd_mamifero.conducta_reproductiva
    ws.cell(row=cont,column=52).value = bd_mamifero.conducta_social
    ws.cell(row=cont,column=53).value = bd_mamifero.especie_segun_ambiente
    ws.cell(row=cont,column=54).value = bd_mamifero.tipo_locomocion
    ws.cell(row=cont,column=55).value = bd_mamifero.tipo_habitat
    ws.cell(row=cont,column=56).value = bd_mamifero.micro_habitat
    ws.cell(row=cont,column=57).value = bd_mamifero.grupo_trofico
    ws.cell(row=cont,column=58).value = bd_mamifero.registro_dieta
    ws.cell(row=cont,column=59).value = bd_mamifero.especie_endemica
    ws.cell(row=cont,column=60).value = bd_mamifero.nivel_endemismo
    ws.cell(row=cont,column=61).value = bd_mamifero.zona_endemismo
    ws.cell(row=cont,column=62).value = bd_mamifero.DSO042014AG
    ws.cell(row=cont,column=63).value = bd_mamifero.IUCN
    ws.cell(row=cont,column=64).value = bd_mamifero.CITES
    ws.cell(row=cont,column=65).value = bd_mamifero.uso_potencial
    ws.cell(row=cont,column=66).value = bd_mamifero.observaciones_sobre_muestra
    ws.cell(row=cont,column=67).value = bd_mamifero.temperie
    ws.cell(row=cont,column=68).value = bd_mamifero.actividad_humana
    ws.cell(row=cont,column=69).value = bd_mamifero.estado_conservacion_habitat
    ws.cell(row=cont,column=70).value = bd_mamifero.colector_registro
    ws.cell(row=cont,column=71).value = bd_mamifero.codigo_colecta
    ws.cell(row=cont,column=72).value = bd_mamifero.tipo_muestra
    ws.cell(row=cont,column=73).value = bd_mamifero.estado_muestra
    ws.cell(row=cont,column=74).value = bd_mamifero.tipo_preparacion_muestra
    ws.cell(row=cont,column=75).value = bd_mamifero.institucion_deposito
    ws.cell(row=cont,column=76).value = bd_mamifero.codigo_fotos
    ws.cell(row=cont,column=77).value = bd_mamifero.observaciones

    nombre_archivo ="BD_Mamiferos_unico.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response


#Coordenadas_mamiferos
def export_coordenadas_reporte_mamiferos(request):

    Datos = Coordenadas_reporte_BD_Mamiferos.objects.all()

    wb = Workbook()
    ws = wb.active

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble
    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    bor = Border(top=double, left=thin, right=thin, bottom=double)
    fillfg = PatternFill("solid", fgColor="00969696")
    fil = fill = GradientFill(stop=("000000", "FFFFFF"))
    font_cell = Font(b=True, color="000000")
    aling = Alignment(horizontal="center", vertical="center")

    ws['E1'] = 'Coordenadas UTM (WGS84)'

    ws.merge_cells('E1:F1')
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:C2')
    ws.merge_cells('D1:D2')
    ws.merge_cells('G1:G2')

    #ws['A1'] = 'Usuario'
    ws['A1'] = 'Sector'
    ws['B1'] = 'Unidad de vegetación'
    ws['C1'] = 'Estacion de muestreo'
    ws['D1'] = 'Unidad de muestreo'
    ws['E2'] = 'Este'
    ws['F2'] = 'Norte'
    ws['G1'] = 'Altitud'

    cont=3

    abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
                       'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ',
                       'BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR','BS','BT','BU','BV','BW','BX','BY','BZ']

    for col in range (1,8):
        cd = ws.cell(row=1,column=col)
        cd.border = bor
        cd.fill = fillfg
        cd.font  = font_cell
        cd.alignment = aling
        ws.column_dimensions[abc[col-1]].width=20.0

    for col in range(1, 8):
            for row in range(1,len(Datos)+1):
                cd = ws.cell(row=row+2,column=col)
                cd.border = bor


    for coordenada in Datos:
        ws.cell(row=cont,column=1).value = coordenada.sector
        ws.cell(row=cont,column=2).value = coordenada.unidad_vegetacion
        ws.cell(row=cont,column=3).value = coordenada.estacion_muestreo
        ws.cell(row=cont,column=4).value = coordenada.unidad_muestreo
        ws.cell(row=cont,column=5).value = coordenada.coordenadas_utm_wgs84.este
        ws.cell(row=cont,column=6).value = coordenada.coordenadas_utm_wgs84.norte
        ws.cell(row=cont,column=7).value = coordenada.altitud

        cont = cont + 1

    nombre_archivo ="Reporte_coordenada_mamiferos.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

#Cordenadas Mamifreos Unico
def export_coordenadas_reporte_mamiferos_unic(request,id):

    coordenada = Coordenadas_reporte_BD_Mamiferos.objects.get(id=id)

    wb = Workbook()
    ws = wb.active

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble
    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    bor = Border(top=double, left=thin, right=thin, bottom=double)
    fillfg = PatternFill("solid", fgColor="00969696")
    fil = fill = GradientFill(stop=("000000", "FFFFFF"))
    font_cell = Font(b=True, color="000000")
    aling = Alignment(horizontal="center", vertical="center")

    ws['E1'] = 'Coordenadas UTM (WGS84)'

    ws.merge_cells('E1:F1')
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:C2')
    ws.merge_cells('D1:D2')
    ws.merge_cells('G1:G2')

    #ws['A1'] = 'Usuario'
    ws['A1'] = 'Sector'
    ws['B1'] = 'Unidad de vegetación'
    ws['C1'] = 'Estacion de muestreo'
    ws['D1'] = 'Unidad de muestreo'
    ws['E2'] = 'Este'
    ws['F2'] = 'Norte'
    ws['G1'] = 'Altitud'

    cont=3

    abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
                       'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ',
                       'BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR','BS','BT','BU','BV','BW','BX','BY','BZ']


    for col in range (1, 8):
            cd = ws.cell(row=1,column=col)
            cd.border = bor
            cd.fill = fillfg
            cd.font  = font_cell
            cd.alignment = aling
            ws.column_dimensions[abc[col-1]].width=20.0

    for col in range(1, 8):
        cd = ws.cell(3,column=col)
        cd.border = bor

    ws.cell(row=cont,column=1).value = coordenada.sector
    ws.cell(row=cont,column=2).value = coordenada.unidad_vegetacion
    ws.cell(row=cont,column=3).value = coordenada.estacion_muestreo
    ws.cell(row=cont,column=4).value = coordenada.unidad_muestreo
    ws.cell(row=cont,column=5).value = coordenada.coordenadas_utm_wgs84.este
    ws.cell(row=cont,column=6).value = coordenada.coordenadas_utm_wgs84.norte
    ws.cell(row=cont,column=7).value = coordenada.altitud

    nombre_archivo ="Reporte_coordenada_mamiferos_unico.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

#Esfuerzo_Muestreo_Mamiferos
def export_esfuerzo_muestreo_mamifero(request):


    Datos = Esfuerzo_Muestreo_BD_Mamiferos.objects.all()

    wb = Workbook()
    ws = wb.active

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble
    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    bor = Border(top=double, left=thin, right=thin, bottom=double)
    fillfg = PatternFill("solid", fgColor="FFAF87")
    fil = fill = GradientFill(stop=("000000", "FFFFFF"))
    font_cell = Font(b=True, color="000000")
    aling = Alignment(horizontal="center", vertical="center")

    ws['A1'] = 'Fecha'
    ws['B1'] = 'Grupo'
    ws['C1'] = 'Temporada'
    ws['D1'] = 'Unidad de vegetación'
    ws['E1'] = 'Estacion de muestreo'
    ws['F1'] = 'Unidad de muestreo'
    ws['G1'] = 'Trampas noche'
    ws['H1'] = 'No. Total Individuos'
    ws['I1'] = 'Especie 1'
    ws['J1'] = 'Especie 2'
    ws['K1'] = 'Especie 3'
    ws['L1'] = 'Especie 4'
    ws['M1'] = 'Clima'
    ws['N1'] = 'Grupo'
    ws['O1'] = 'Estacion de muestreo'
    ws['P1'] = 'Unidad de Muestreo (Tramsecto)'
    ws['Q1'] = 'Coordenadas UTM'
    ws['R1'] = 'Coordenadas UTM'
    ws['S1'] = 'Elevación (m)'
    ws['T1'] = 'Coordenadas UTM'
    ws['U1'] = 'Coordenadas UTM'
    ws['V1'] = 'Elevación (m)'
    ws['W1'] = 'm recoridos'
    ws['X1'] = 'Coordenadas UTM'
    ws['Y1'] = 'Coordenadas UTM'

    ws['Z1'] = 'Elevación (m)'
    ws['AA1'] = 'Coordenadas UTM'
    ws['AB1'] = 'Coordenadas UTM'
    ws['AC1'] = 'Elevación (m)'
    ws['AD1'] = 'Coordenadas UTM'
    ws['AE1'] = 'Coordenadas UTM'
    ws['AF1'] = 'Elevación (m)'
    ws['AG1'] = 'Coordenadas UTM'
    ws['AH1'] = 'Coordenadas UTM'
    ws['AI1'] = 'Elevación (m)'

    ws['AJ1'] = 'Unidad de muestreo'
    ws['AK1'] = 'Mamíferos Menores'
    ws['AL1'] = 'Mamíferos Mayores'

    cont=2

    abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
                       'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ',
                       'BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR','BS','BT','BU','BV','BW','BX','BY','BZ']

    for col in range (1,39):
        cd = ws.cell(row=1,column=col)
        cd.border = bor
        cd.fill = fillfg
        cd.font  = font_cell
        cd.alignment = aling
        ws.column_dimensions[abc[col-1]].width=20.0

    for col in range(1, 39):
            for row in range(1,len(Datos)+1):
                cd = ws.cell(row=row+2,column=col)
                cd.border = bor

    for coordenada in Datos:
        ws.cell(row=cont,column=1, value = coordenada.fecha).number_format = numbers.FORMAT_DATE_XLSX15
        ws.cell(row=cont,column=2).value = coordenada.grupo1
        ws.cell(row=cont,column=3).value = coordenada.temporada
        ws.cell(row=cont,column=4).value = coordenada.unidad_vegetacion
        ws.cell(row=cont,column=5).value = coordenada.estacion_muestreo1
        ws.cell(row=cont,column=6).value = coordenada.unidad_muestreo
        ws.cell(row=cont,column=7).value = coordenada.trampas_noche
        ws.cell(row=cont,column=8).value = coordenada.numero_total_individuos
        ws.cell(row=cont,column=9).value = coordenada.especie1
        ws.cell(row=cont,column=10).value = coordenada.especie2
        ws.cell(row=cont,column=11).value = coordenada.especie3
        ws.cell(row=cont,column=12).value = coordenada.especie4
        ws.cell(row=cont,column=13).value = coordenada.clima
        ws.cell(row=cont,column=14).value = coordenada.grupo
        ws.cell(row=cont,column=15).value = coordenada.estacion_muestreo
        ws.cell(row=cont,column=16).value = coordenada.unidad_muestreo_transecto
        ws.cell(row=cont,column=17).value = coordenada.coordenadas_utm
        ws.cell(row=cont,column=18).value = coordenada.coordenadas2_utm
        ws.cell(row=cont,column=19).value = coordenada.elevacion_metros
        ws.cell(row=cont,column=20).value = coordenada.coordenadas3_utm
        ws.cell(row=cont,column=21).value = coordenada.coordenadas4_utm
        ws.cell(row=cont,column=22).value = coordenada.elevacion2_metros
        ws.cell(row=cont,column=23).value = coordenada.metros_recorridos
        ws.cell(row=cont,column=24).value = coordenada.coordenadas5_utm
        ws.cell(row=cont,column=25).value = coordenada.coordenadas6_utm
        ws.cell(row=cont,column=26).value = coordenada.elevacion3_metros
        ws.cell(row=cont,column=27).value = coordenada.coordenadas7_utm
        ws.cell(row=cont,column=28).value = coordenada.coordenadas8_utm
        ws.cell(row=cont,column=29).value = coordenada.elevacion4_metros
        ws.cell(row=cont,column=30).value = coordenada.coordenadas9_utm
        ws.cell(row=cont,column=31).value = coordenada.coordenadas10_utm
        ws.cell(row=cont,column=32).value = coordenada.elevacion5_metros
        ws.cell(row=cont,column=33).value = coordenada.coordenadas11_utm
        ws.cell(row=cont,column=34).value = coordenada.coordenadas12_utm
        ws.cell(row=cont,column=35).value = coordenada.elevacion6_metros
        ws.cell(row=cont,column=36).value = coordenada.unidad
        ws.cell(row=cont,column=37).value = coordenada.mamiferos_menores
        ws.cell(row=cont,column=38).value = coordenada.mamiferos_mayores
        cont = cont + 1

    nombre_archivo ="Esfuerzo_de_muestreo_mamiferos.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

#Esfuerzo Muestreo Unico

def export_esfuerzo_muestreo_mamiferos_unico(request, id):

    coordenada = Esfuerzo_Muestreo_BD_Mamiferos.objects.get(id=id)

    wb = Workbook()
    ws = wb.active
    #ws['J1'] = 'Georreferenciación WGS 84-18L'

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble
    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    bor = Border(top=double, left=thin, right=thin, bottom=double)
    fillfg = PatternFill("solid", fgColor="FFAF87")
    fil = fill = GradientFill(stop=("000000", "FFFFFF"))
    font_cell = Font(b=True, color="000000")
    aling = Alignment(horizontal="center", vertical="center")

    #ws['A1'] = 'Usuario'
    ws['A1'] = 'Fecha'
    ws['B1'] = 'Grupo'
    ws['C1'] = 'Temporada'
    ws['D1'] = 'Unidad de vegetación'
    ws['E1'] = 'Estacion de muestreo'
    ws['F1'] = 'Unidad de muestreo'
    ws['G1'] = 'Trampas noche'
    ws['H1'] = 'No. Total Individuos'
    ws['I1'] = 'Especie 1'
    ws['J1'] = 'Especie 2'
    ws['K1'] = 'Especie 3'
    ws['L1'] = 'Especie 4'
    ws['M1'] = 'Clima'
    ws['N1'] = 'Grupo'
    ws['O1'] = 'Estacion de muestreo'
    ws['P1'] = 'Unidad de Muestreo (Tramsecto)'
    ws['Q1'] = 'Coordenadas UTM'
    ws['R1'] = 'Coordenadas UTM'
    ws['S1'] = 'Elevación (m)'
    ws['T1'] = 'Coordenadas UTM'
    ws['U1'] = 'Coordenadas UTM'
    ws['V1'] = 'Elevación (m)'
    ws['W1'] = 'm recoridos'
    ws['X1'] = 'Coordenadas UTM'
    ws['Y1'] = 'Coordenadas UTM'

    ws['Z1'] = 'Elevación (m)'
    ws['AA1'] = 'Coordenadas UTM'
    ws['AB1'] = 'Coordenadas UTM'
    ws['AC1'] = 'Elevación (m)'
    ws['AD1'] = 'Coordenadas UTM'
    ws['AE1'] = 'Coordenadas UTM'
    ws['AF1'] = 'Elevación (m)'
    ws['AG1'] = 'Coordenadas UTM'
    ws['AH1'] = 'Coordenadas UTM'
    ws['AI1'] = 'Elevación (m)'

    ws['AJ1'] = 'Unidad de muestreo'
    ws['AK1'] = 'Mamíferos Menores'
    ws['AL1'] = 'Mamíferos Mayores'

    cont=2

    abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
                       'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ',
                       'BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR','BS','BT','BU','BV','BW','BX','BY','BZ']


    for col in range (1, 39):
            cd = ws.cell(row=1,column=col)
            cd.border = bor
            cd.fill = fillfg
            cd.font  = font_cell
            cd.alignment = aling
            ws.column_dimensions[abc[col-1]].width=20.0

    for col in range(1, 39):
        cd = ws.cell(2,column=col)
        cd.border = bor


    ws.cell(row=cont,column=1,value = coordenada.fecha).number_format = numbers.FORMAT_DATE_XLSX15
    ws.cell(row=cont,column=2).value = coordenada.grupo1
    ws.cell(row=cont,column=3).value = coordenada.temporada
    ws.cell(row=cont,column=4).value = coordenada.unidad_vegetacion
    ws.cell(row=cont,column=5).value = coordenada.estacion_muestreo1
    ws.cell(row=cont,column=6).value = coordenada.unidad_muestreo
    ws.cell(row=cont,column=7).value = coordenada.trampas_noche
    ws.cell(row=cont,column=8).value = coordenada.numero_total_individuos
    ws.cell(row=cont,column=9).value = coordenada.especie1
    ws.cell(row=cont,column=10).value = coordenada.especie2
    ws.cell(row=cont,column=11).value = coordenada.especie3
    ws.cell(row=cont,column=12).value = coordenada.especie4
    ws.cell(row=cont,column=13).value = coordenada.clima
    ws.cell(row=cont,column=14).value = coordenada.grupo
    ws.cell(row=cont,column=15).value = coordenada.estacion_muestreo
    ws.cell(row=cont,column=16).value = coordenada.unidad_muestreo_transecto
    ws.cell(row=cont,column=17).value = coordenada.coordenadas_utm
    ws.cell(row=cont,column=18).value = coordenada.coordenadas2_utm
    ws.cell(row=cont,column=19).value = coordenada.elevacion_metros
    ws.cell(row=cont,column=20).value = coordenada.coordenadas3_utm
    ws.cell(row=cont,column=21).value = coordenada.coordenadas4_utm
    ws.cell(row=cont,column=22).value = coordenada.elevacion2_metros
    ws.cell(row=cont,column=23).value = coordenada.metros_recorridos
    ws.cell(row=cont,column=24).value = coordenada.coordenadas5_utm
    ws.cell(row=cont,column=25).value = coordenada.coordenadas6_utm
    ws.cell(row=cont,column=26).value = coordenada.elevacion3_metros
    ws.cell(row=cont,column=27).value = coordenada.coordenadas7_utm
    ws.cell(row=cont,column=28).value = coordenada.coordenadas8_utm
    ws.cell(row=cont,column=29).value = coordenada.elevacion4_metros
    ws.cell(row=cont,column=30).value = coordenada.coordenadas9_utm
    ws.cell(row=cont,column=31).value = coordenada.coordenadas10_utm
    ws.cell(row=cont,column=32).value = coordenada.elevacion5_metros
    ws.cell(row=cont,column=33).value = coordenada.coordenadas11_utm
    ws.cell(row=cont,column=34).value = coordenada.coordenadas12_utm
    ws.cell(row=cont,column=35).value = coordenada.elevacion6_metros
    ws.cell(row=cont,column=36).value = coordenada.unidad
    ws.cell(row=cont,column=37).value = coordenada.mamiferos_menores
    ws.cell(row=cont,column=38).value = coordenada.mamiferos_mayores


    nombre_archivo ="Esfuerzo_de_muestreo_mamiferos.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

#Bitacora_Mamiferos
def export_bitacora_mamiferos(request):

    Datos = bitacora_BD_Mamiferos.objects.all()

    wb = Workbook()
    ws = wb.active

    #ws['J1'] = 'Georreferenciación WGS 84-18L'

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble
    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    bor = Border(top=double, left=thin, right=thin, bottom=double)
    fillfg = PatternFill("solid", fgColor="EEEEEE")
    fil = fill = GradientFill(stop=("000000", "FFFFFF"))
    font_cell = Font(b=True, color="000000")
    aling = Alignment(horizontal="center", vertical="center")


    #ws.merge_cells('J1:P1')
    ws.merge_cells('A1:A2')
    ws.merge_cells('B1:B2')
    ws.merge_cells('C1:C2')

    #ws['A1'] = 'Usuario'
    ws['A1'] = 'Dia'
    ws['B1'] = 'Fecha'
    ws['C1'] = 'Actividades'

    cont = 3


    abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
                       'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ',
                       'BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR','BS','BT','BU','BV','BW','BX','BY','BZ']

    for col in range (1,4):
        cd = ws.cell(row=1,column=col)
        cd.border = bor
        cd.fill = fillfg
        cd.font  = font_cell
        cd.alignment = aling
        ws.column_dimensions[abc[col-1]].width=20.0

    for col in range(1, 4):
            for row in range(1,len(Datos)+1):
                cd = ws.cell(row=row+2,column=col)
                cd.border = bor


    for coordenada in Datos:
            ws.cell(row=cont,column=1).value = coordenada.dia
            ws.cell(row=cont,column=2,value = coordenada.fecha).number_format = numbers.FORMAT_DATE_XLSX15
            ws.cell(row=cont,column=3).value = coordenada.actividades

            cont = cont + 1

    nombre_archivo ="Bitacora_Mamiferos.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

#Bitacora_Mamiferos_Unico
def export_bitacora_mamiferos_unico(request,id):
    coordenada = bitacora_BD_Mamiferos.objects.get(id=id)

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble
    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    bor = Border(top=double, left=thin, right=thin, bottom=double)
    fillfg = PatternFill("solid", fgColor="EEEEEE")
    fil = fill = GradientFill(stop=("000000", "FFFFFF"))
    font_cell = Font(b=True, color="000000")
    aling = Alignment(horizontal="center", vertical="center")


    wb = Workbook()
    ws = wb.active

    #ws['A1'] = 'Usuario'
    ws['A1'] = 'Dia'
    ws['B1'] = 'Fecha'
    ws['C1'] = 'Actividades'


    cont = 2

    abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
                       'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ',
                       'BA','BB','BC','BD','BE','BF','BG','BH','BI','BJ','BK','BL','BM','BN','BO','BP','BQ','BR','BS','BT','BU','BV','BW','BX','BY','BZ']


    for col in range (1, 4):
            cd = ws.cell(row=1,column=col)
            cd.border = bor
            cd.fill = fillfg
            cd.font  = font_cell
            cd.alignment = aling
            ws.column_dimensions[abc[col-1]].width=20.0

    for col in range(1, 4):
        cd = ws.cell(2,column=col)
        cd.border = bor

    ws.cell(row=cont,column=1).value = coordenada.dia
    ws.cell(row=cont,column=2,value = coordenada.fecha).number_format = numbers.FORMAT_DATE_XLSX15
    ws.cell(row=cont,column=3).value = coordenada.actividades

    nombre_archivo ="Bitacora_Mamiferos.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

# ============================== SECCION VEGETACION =====================================

#======================================
# VEGETACION HOJA3
def export_vegetacion_hoja3(request):

    datos = Hoja3_BD_Vegetacion_2018.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja3"

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble

    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    borde_celda_1 = Border(top=double, left=thin, right=thin, bottom=double)
    borde_celda_2 = Border(top=thin, left=thin, right=thin, bottom=thin)
    fondo_celda_cabeza = PatternFill("solid", fgColor="2EFF00")
    fondo_texto_cabeza = Font(b=True, color="000000")
    alineacion_celda = Alignment(horizontal="center", vertical="center")
    orientacion_alineacion_celda = Alignment(horizontal="center", vertical="center", textRotation=90)


    ws['A1'] = 'Especies'
    ws['B1'] = 'D.S. 043-2006-AG'
    ws['C1'] = 'Endémico'
    ws['D1'] = 'IUCN'
    ws['E1'] = 'CITES'

    abc = ['A','B','C','D','E']

    for col in range (1,6):
        cd = ws.cell(row=1,column=col)
        cd.border = borde_celda_1
        cd.fill = fondo_celda_cabeza
        cd.font = fondo_texto_cabeza
        cd.alignment = alineacion_celda
        ws.column_dimensions[abc[col-1]].width=18.0
        ws.row_dimensions[1].height = 30

    for col in range(1, 6):
        for row in range(1,len(datos)+1):
            cd = ws.cell(row=row+1,column=col)
            cd.border = borde_celda_2

    cont=2

    for dato in datos:
        ws.cell(row=cont,column=1).value = dato.especies
        ws.cell(row=cont,column=2).value = dato.DS0432006AG
        ws.cell(row=cont,column=3).value = dato.endemico
        ws.cell(row=cont,column=4).value = dato.IUCN
        ws.cell(row=cont,column=5).value = dato.CITES

        cont = cont + 1

    nombre_archivo ="Vegetacion Hoja3.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

# VEGETACION HOJA3 UNICO
def export_vegetacion_hoja3_unic(request, id):

    dato = Hoja3_BD_Vegetacion_2018.objects.get(id=id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja3"

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble

    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    borde_celda_1 = Border(top=double, left=thin, right=thin, bottom=double)
    borde_celda_2 = Border(top=thin, left=thin, right=thin, bottom=thin)
    fondo_celda_cabeza = PatternFill("solid", fgColor="2EFF00")
    fondo_texto_cabeza = Font(b=True, color="000000")
    alineacion_celda = Alignment(horizontal="center", vertical="center")
    orientacion_alineacion_celda = Alignment(horizontal="center", vertical="center", textRotation=90)


    ws['A1'] = 'Especies'
    ws['B1'] = 'D.S. 043-2006-AG'
    ws['C1'] = 'Endémico'
    ws['D1'] = 'IUCN'
    ws['E1'] = 'CITES'

    abc = ['A','B','C','D','E']

    for col in range (1,6):
        cd = ws.cell(row=1,column=col)
        cd.border = borde_celda_1
        cd.fill = fondo_celda_cabeza
        cd.font = fondo_texto_cabeza
        cd.alignment = alineacion_celda
        ws.column_dimensions[abc[col-1]].width=18.0
        ws.row_dimensions[1].height = 30

    for col in range(1, 6):
        cd = ws.cell(2, column=col)
        cd.border = borde_celda_2

    cont=2

    ws.cell(row=cont,column=1).value = dato.especies
    ws.cell(row=cont,column=2).value = dato.DS0432006AG
    ws.cell(row=cont,column=3).value = dato.endemico
    ws.cell(row=cont,column=4).value = dato.IUCN
    ws.cell(row=cont,column=5).value = dato.CITES

    nombre_archivo ="Reporte_Vegetacion_Hoja3_Unico.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response
#======================================

#======================================
# VEGETACION BASE DE DATOS
def export_vegetacion_bd(request):

    datos = Base_Datos_Vegetacion_2018.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Base de Datos"

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble

    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    borde_celda_1 = Border(top=double, left=thin, right=thin, bottom=double)
    borde_celda_2 = Border(top=thin, left=thin, right=thin, bottom=thin)
    fondo_celda_cabeza = PatternFill("solid", fgColor="C4D79B")
    fondo_celda_especial = PatternFill("solid", fgColor="FFFF00")
    fondo_texto_cabeza = Font(b=True, color="000000")
    alineacion_celda = Alignment(horizontal="center", vertical="center")
    orientacion_alineacion_celda = Alignment(horizontal="center", vertical="center", textRotation=90)


    ws['A1'] = 'N°'
    ws['B1'] = 'Disciplina'
    ws['C1'] = 'Código de proyecto'
    ws['D1'] = 'Fecha de registro'
    ws['E1'] = 'Temporada'
    ws['F1'] = 'Unidad de vegetación'
    ws['G1'] = 'Estación de muestreo'
    ws['H1'] = 'Unidad de muestreo'
    ws['I1'] = 'Subunidad de muestreo'
    ws['J1'] = 'Método de muestreo'
    ws['K1'] = 'Tamaño de la unidad de muestreo'
    ws['L1'] = 'División'
    ws['M1'] = 'Orden'
    ws['N1'] = 'Familia'
    ws['O1'] = 'Especie'

    ws['P1'] = 'Nombre común'
    ws['Q1'] = 'Nombre local'

    ws['R1'] = 'Fuente de clasificación taxonomica'
    ws['S1'] = ' Cobertura vegetal (%)'
    ws['T1'] = 'X (NORTE)'
    ws['U1'] = 'Y (ESTE)'
    ws['V1'] = 'Proyección'
    ws['W1'] = 'Altitud (msnm)'
    ws['X1'] = 'Datum'

    ws['Y1'] = 'Número de individuos'
    ws['Z1'] = 'DAP (cm)'
    ws['AA1'] = 'Altura total (m)'

    ws['AB1'] = 'Estado de conservación del hábitat'
    ws['AC1'] = 'Fenología'
    ws['AD1'] = 'Habito'
    ws['AE1'] = 'Colector de registro'
    ws['AF1'] = 'Tipo de Evaluación'

    ws['AG1'] = 'Uso'
    ws['AH1'] = 'Tipo de uso'

    ws['AI1'] = 'D.S. 043-2006-AG'
    ws['AJ1'] = 'Endemico'
    ws['AK1'] = 'IUCN'
    ws['AL1'] = 'CITES'
    ws['AM1'] = 'FOTO'
    ws['AN1'] = 'TIPO DE GANADO OBSERVADO '
    ws['AO1'] = 'Observación'

    abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z',
           'AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO']

    for col in range (1,42):
        cd = ws.cell(row=1,column=col)
        cd.border = borde_celda_1
        if (col == 16 or col == 17 or col == 25 or col == 26 or col == 27 or col == 33 or col == 34):
            cd.fill = fondo_celda_especial
        else:
            cd.fill = fondo_celda_cabeza

        cd.font = fondo_texto_cabeza
        if (col == 1):
            cd.alignment = alineacion_celda
        else:
            cd.alignment = orientacion_alineacion_celda

        ws.column_dimensions[abc[col-1]].width=10.0
        ws.row_dimensions[1].height = 100

    for col in range(1, 42):
        for row in range(1,len(datos)+1):
            cd = ws.cell(row=row+1,column=col)
            cd.border = borde_celda_2

    cont=2

    for dato in datos:
        ws.cell(row=cont,column=1).value = dato.numero
        ws.cell(row=cont,column=2).value = dato.disciplina
        ws.cell(row=cont,column=3).value = dato.codigo_proyecto
        ws.cell(row=cont,column=4, value = dato.fecha_registro).number_format = numbers.FORMAT_DATE_XLSX15
        ws.cell(row=cont,column=5).value = dato.temporada
        ws.cell(row=cont,column=6).value = dato.unidad_vegetacion
        ws.cell(row=cont,column=7).value = dato.estacion_muestreo
        ws.cell(row=cont,column=8).value = dato.unidad_muestreo
        ws.cell(row=cont,column=9).value = dato.subunidad_muestreo
        ws.cell(row=cont,column=10).value = dato.metodo_muestreo
        ws.cell(row=cont,column=11).value = dato.tamaño_unidad_muestreo
        ws.cell(row=cont,column=12).value = dato.division
        ws.cell(row=cont,column=13).value = dato.orden
        ws.cell(row=cont,column=14).value = dato.familia
        ws.cell(row=cont,column=15).value = dato.especie
        ws.cell(row=cont,column=16).value = dato.nombre_comun
        ws.cell(row=cont,column=17).value = dato.nombre_local
        ws.cell(row=cont,column=18).value = dato.fuente_clasificacion_taxonomica
        ws.cell(row=cont,column=19).value = dato.cobertura_vegetal
        ws.cell(row=cont,column=20).value = dato.x_norte
        ws.cell(row=cont,column=21).value = dato.y_norte
        ws.cell(row=cont,column=22).value = dato.proyeccion
        ws.cell(row=cont,column=23).value = dato.altitud_msnm
        ws.cell(row=cont,column=24).value = dato.datum
        ws.cell(row=cont,column=25).value = dato.numero_individuos
        ws.cell(row=cont,column=26).value = dato.DAP_cm
        ws.cell(row=cont,column=27).value = dato.altura_total
        ws.cell(row=cont,column=28).value = dato.estado_conservacion_habitat
        ws.cell(row=cont,column=29).value = dato.fenologia
        ws.cell(row=cont,column=30).value = dato.habito
        ws.cell(row=cont,column=31).value = dato.colector_registro
        ws.cell(row=cont,column=32).value = dato.tipo_evaluacion
        ws.cell(row=cont,column=33).value = dato.uso
        ws.cell(row=cont,column=34).value = dato.tipo_uso
        ws.cell(row=cont,column=35).value = dato.DS0432006AG
        ws.cell(row=cont,column=36).value = dato.endemico
        ws.cell(row=cont,column=37).value = dato.IUCN
        ws.cell(row=cont,column=38).value = dato.CITES
        ws.cell(row=cont,column=39).value = dato.FOTO
        ws.cell(row=cont,column=40).value = dato.tipo_ganado_observado
        ws.cell(row=cont,column=41).value = dato.observacion

        cont = cont + 1

    nombre_archivo ="Vegetacion BD.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

# VEGETACION BASE DE DATOS UNICO
def export_basedatos_vegetacion_unico(request,id):
    dato = Base_Datos_Vegetacion_2018.objects.get(id=id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Base de Datos"



    cont=1

    ws.cell(row=cont,column=1).value = dato.numero
    ws.cell(row=cont,column=2).value = dato.disciplina
    ws.cell(row=cont,column=3).value = dato.codigo_proyecto
    ws.cell(row=cont,column=4, value = dato.fecha_registro).number_format = numbers.FORMAT_DATE_XLSX15
    ws.cell(row=cont,column=5).value = dato.temporada
    ws.cell(row=cont,column=6).value = dato.unidad_vegetacion
    ws.cell(row=cont,column=7).value = dato.estacion_muestreo
    ws.cell(row=cont,column=8).value = dato.unidad_muestreo
    ws.cell(row=cont,column=9).value = dato.subunidad_muestreo
    ws.cell(row=cont,column=10).value = dato.metodo_muestreo
    ws.cell(row=cont,column=11).value = dato.tamaño_unidad_muestreo
    ws.cell(row=cont,column=12).value = dato.division
    ws.cell(row=cont,column=13).value = dato.orden
    ws.cell(row=cont,column=14).value = dato.familia
    ws.cell(row=cont,column=15).value = dato.especie
    ws.cell(row=cont,column=16).value = dato.nombre_comun
    ws.cell(row=cont,column=17).value = dato.nombre_local
    ws.cell(row=cont,column=18).value = dato.fuente_clasificacion_taxonomica
    ws.cell(row=cont,column=19).value = dato.cobertura_vegetal
    ws.cell(row=cont,column=20).value = dato.x_norte
    ws.cell(row=cont,column=21).value = dato.y_norte
    ws.cell(row=cont,column=22).value = dato.proyeccion
    ws.cell(row=cont,column=23).value = dato.altitud_msnm
    ws.cell(row=cont,column=24).value = dato.datum
    ws.cell(row=cont,column=25).value = dato.numero_individuos
    ws.cell(row=cont,column=26).value = dato.DAP_cm
    ws.cell(row=cont,column=27).value = dato.altura_total
    ws.cell(row=cont,column=28).value = dato.estado_conservacion_habitat
    ws.cell(row=cont,column=29).value = dato.fenologia
    ws.cell(row=cont,column=30).value = dato.habito
    ws.cell(row=cont,column=31).value = dato.colector_registro
    ws.cell(row=cont,column=32).value = dato.tipo_evaluacion
    ws.cell(row=cont,column=33).value = dato.uso
    ws.cell(row=cont,column=34).value = dato.tipo_uso
    ws.cell(row=cont,column=35).value = dato.DS0432006AG
    ws.cell(row=cont,column=36).value = dato.endemico
    ws.cell(row=cont,column=37).value = dato.IUCN
    ws.cell(row=cont,column=38).value = dato.CITES
    ws.cell(row=cont,column=39).value = dato.FOTO
    ws.cell(row=cont,column=40).value = dato.tipo_ganado_observado
    ws.cell(row=cont,column=41).value = dato.observacion



    nombre_archivo ="Vegetacion BD.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response


#======================================

#======================================
# VEGETACION ESFEURZO DE MUESTREO
def export_vegetacion_esfuerzo(request):

    datos = Esfuerzo_Muestreo_BD_Vegetacion_2018.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Esfuerzo de muestreo"

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble

    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    borde_celda_1 = Border(top=double, left=thin, right=thin, bottom=double)
    borde_celda_2 = Border(top=thin, left=thin, right=thin, bottom=thin)
    fondo_celda_cabeza = PatternFill("solid", fgColor="C4D79B")
    fondo_celda_especial = PatternFill("solid", fgColor="FFFF00")
    fondo_texto_cabeza = Font(b=True, color="000000")
    alineacion_celda = Alignment(horizontal="center", vertical="center")
    orientacion_alineacion_celda = Alignment(horizontal="center", vertical="center", textRotation=90)


    ws['A1'] = 'Código de proyecto'
    ws['B1'] = 'Fecha'
    ws['C1'] = 'Temporada'
    ws['D1'] = 'Unidad de vegetación'
    ws['E1'] = 'Sector'
    ws['F1'] = 'Estación de muestreo'
    ws['G1'] = 'Unidad de muestreo'
    ws['H1'] = 'Esfuerzo de muestreo'
    ws['I1'] = 'Unidad del esfuerzo de muestreo'
    ws['J1'] = 'Tamaño de la unidad'

    abc = ['A','B','C','D','E','F','G','H','I','J']

    for col in range (1,11):
        cd = ws.cell(row=1,column=col)
        cd.border = borde_celda_1
        cd.fill = fondo_celda_cabeza
        cd.font = fondo_texto_cabeza
        cd.alignment = alineacion_celda
        ws.column_dimensions[abc[col-1]].width = 19
        ws.row_dimensions[1].height = 30

    for col in range(1, 11):
        for row in range(1,len(datos)+1):
            cd = ws.cell(row=row+1,column=col)
            cd.border = borde_celda_2

    cont=2

    for dato in datos:
        ws.cell(row=cont,column=1).value = dato.codigo_proyecto
        ws.cell(row=cont,column=2, value = dato.fecha).number_format = numbers.FORMAT_DATE_XLSX15
        ws.cell(row=cont,column=3).value = dato.temporada
        ws.cell(row=cont,column=4).value = dato.unidad_vegetacion
        ws.cell(row=cont,column=5).value = dato.sector
        ws.cell(row=cont,column=6).value = dato.estacion_muestreo
        ws.cell(row=cont,column=7).value = dato.unidad_muestreo
        ws.cell(row=cont,column=8).value = dato.esfuerzo_muestreo
        ws.cell(row=cont,column=9).value = dato.unidad_esfuerzo_muestreo
        ws.cell(row=cont,column=10).value = dato.tamaño_unidad_metros

        cont = cont + 1

    nombre_archivo ="Vegetacion Esfuerzo de Muestreo.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

# VEGETACION ESFEURZO DE MUESTREO UNICO
def export_vegetacion_esfuerzo_unic(request, id):

    dato = Esfuerzo_Muestreo_BD_Vegetacion_2018.objects.get(id=id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Esfuerzo de muestreo"

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble

    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    borde_celda_1 = Border(top=double, left=thin, right=thin, bottom=double)
    borde_celda_2 = Border(top=thin, left=thin, right=thin, bottom=thin)
    fondo_celda_cabeza = PatternFill("solid", fgColor="C4D79B")
    fondo_celda_especial = PatternFill("solid", fgColor="FFFF00")
    fondo_texto_cabeza = Font(b=True, color="000000")
    alineacion_celda = Alignment(horizontal="center", vertical="center")
    orientacion_alineacion_celda = Alignment(horizontal="center", vertical="center", textRotation=90)


    ws['A1'] = 'Código de proyecto'
    ws['B1'] = 'Fecha'
    ws['C1'] = 'Temporada'
    ws['D1'] = 'Unidad de vegetación'
    ws['E1'] = 'Sector'
    ws['F1'] = 'Estación de muestreo'
    ws['G1'] = 'Unidad de muestreo'
    ws['H1'] = 'Esfuerzo de muestreo'
    ws['I1'] = 'Unidad del esfuerzo de muestreo'
    ws['J1'] = 'Tamaño de la unidad'

    abc = ['A','B','C','D','E','F','G','H','I','J']

    for col in range (1,11):
        cd = ws.cell(row=1,column=col)
        cd.border = borde_celda_1
        cd.fill = fondo_celda_cabeza
        cd.font = fondo_texto_cabeza
        cd.alignment = alineacion_celda
        ws.column_dimensions[abc[col-1]].width = 19
        ws.row_dimensions[1].height = 30

    for col in range(1, 11):
        cd = ws.cell(2,column=col)
        cd.border = borde_celda_2

    cont=2

    ws.cell(row=cont,column=1).value = dato.codigo_proyecto
    ws.cell(row=cont,column=2, value = dato.fecha).number_format = numbers.FORMAT_DATE_XLSX15
    ws.cell(row=cont,column=3).value = dato.temporada
    ws.cell(row=cont,column=4).value = dato.unidad_vegetacion
    ws.cell(row=cont,column=5).value = dato.sector
    ws.cell(row=cont,column=6).value = dato.estacion_muestreo
    ws.cell(row=cont,column=7).value = dato.unidad_muestreo
    ws.cell(row=cont,column=8).value = dato.esfuerzo_muestreo
    ws.cell(row=cont,column=9).value = dato.unidad_esfuerzo_muestreo
    ws.cell(row=cont,column=10).value = dato.tamaño_unidad_metros

    nombre_archivo ="Vegetacion_Esfuerzo_de_Muestreo_Unico.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response
#======================================

#======================================
# VEGETACION COMPENSACION AMBIENTAL
def export_vegetacion_compensacion(request):

    datos = Compensacion_Ambiental_Vegetacion_2018.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Compensacion Amb"

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble

    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    borde_celda_1 = Border(top=double, left=thin, right=thin, bottom=double)
    borde_celda_2 = Border(top=thin, left=thin, right=thin, bottom=thin)
    fondo_celda_cabeza = PatternFill("solid", fgColor="C4D79B")
    fondo_celda_especial = PatternFill("solid", fgColor="FFFF00")
    fondo_texto_cabeza = Font(b=True, color="000000")
    alineacion_celda = Alignment(horizontal="center", vertical="center")
    orientacion_alineacion_celda = Alignment(horizontal="center", vertical="center", textRotation=90)


    ws['A1'] = 'Disciplina'
    ws['B1'] = 'Código de proyecto'
    ws['C1'] = 'Fecha de registro'
    ws['D1'] = 'Temporada'
    ws['E1'] = 'Unidad de vegetación'
    ws['F1'] = 'Estación de muestreo'
    ws['G1'] = 'Unidad de muestreo'
    ws['H1'] = 'Subunidad de muestreo'
    ws['I1'] = 'Método de muestreo'
    ws['J1'] = 'Tamaño de la unidad de muestreo'
    ws['K1'] = 'Especie dominante'
    ws['L1'] = 'X (NORTE)'
    ws['M1'] = 'Y (ESTE)'
    ws['N1'] = 'Proyección'
    ws['O1'] = 'Altitud (msnm)'
    ws['P1'] = 'Datum'
    ws['Q1'] = 'Altura total de la muestra (m)'
    ws['R1'] = 'Área de la muestra (l x a)'
    ws['S1'] = 'Profundidad total de calicata (m)'
    ws['T1'] = 'altura de la vegetación '
    ws['U1'] = 'Altura de la materia organica'
    ws['V1'] = 'Estado de conservación del hábitat'
    ws['W1'] = 'Colector de registro'
    ws['X1'] = 'FOTO'
    ws['Y1'] = 'Observación de IMPACTO'

    abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y']

    for col in range (1,26):
        cd = ws.cell(row=1,column=col)
        cd.border = borde_celda_1
        cd.fill = fondo_celda_cabeza
        cd.font = fondo_texto_cabeza
        cd.alignment = orientacion_alineacion_celda
        ws.column_dimensions[abc[col-1]].width=11
        ws.row_dimensions[1].height = 90

    for col in range(1, 26):
        for row in range(1,len(datos)+1):
            cd = ws.cell(row=row+1,column=col)
            cd.border = borde_celda_2

    cont=2

    for dato in datos:
        ws.cell(row=cont,column=1).value = dato.disciplina
        ws.cell(row=cont,column=2).value = dato.codigo_proyecto
        ws.cell(row=cont,column=3, value = dato.fecha_registro).number_format = numbers.FORMAT_DATE_XLSX15
        ws.cell(row=cont,column=4).value = dato.temporada
        ws.cell(row=cont,column=5).value = dato.unidad_vegetacion
        ws.cell(row=cont,column=6).value = dato.estacion_muestreo
        ws.cell(row=cont,column=7).value = dato.unidad_muestreo
        ws.cell(row=cont,column=8).value = dato.subunidad_muestreo
        ws.cell(row=cont,column=9).value = dato.metodo_muestreo
        ws.cell(row=cont,column=10).value = dato.tamaño_unidad_muestreo
        ws.cell(row=cont,column=11).value = dato.especie_dominante
        ws.cell(row=cont,column=12).value = dato.x_norte
        ws.cell(row=cont,column=13).value = dato.y_este
        ws.cell(row=cont,column=14).value = dato.proyeccion
        ws.cell(row=cont,column=15).value = dato.altitud_msnm
        ws.cell(row=cont,column=16).value = dato.datum
        ws.cell(row=cont,column=17).value = dato.altura_total_muestra
        ws.cell(row=cont,column=18).value = dato.area_muestra
        ws.cell(row=cont,column=19).value = dato.profundidad_total_calicata
        ws.cell(row=cont,column=20).value = dato.altura_vegetacion
        ws.cell(row=cont,column=21).value = dato.altura_materia_organica
        ws.cell(row=cont,column=22).value = dato.estado_conservacion_habitat
        ws.cell(row=cont,column=23).value = dato.colector_registro
        ws.cell(row=cont,column=24).value = dato.FOTO
        ws.cell(row=cont,column=25).value = dato.observacion_impacto

        cont = cont + 1

    nombre_archivo ="Vegetacion Compensacion Amb.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

# VEGETACION COMPENSACION AMBIENTAL UNICO
def export_vegetacion_compensacion_unic(request, id):

    dato = Compensacion_Ambiental_Vegetacion_2018.objects.get(id=id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Compensacion Amb"

    #estilos bordes
    thin = Side(border_style="thin", color="000000") #Borde unico
    double = Side(border_style="double", color="000000") #Borde doble

    #Aplicacion de etilos (bordes, bg, bg-graf, font, aling)
    borde_celda_1 = Border(top=double, left=thin, right=thin, bottom=double)
    borde_celda_2 = Border(top=thin, left=thin, right=thin, bottom=thin)
    fondo_celda_cabeza = PatternFill("solid", fgColor="C4D79B")
    fondo_celda_especial = PatternFill("solid", fgColor="FFFF00")
    fondo_texto_cabeza = Font(b=True, color="000000")
    alineacion_celda = Alignment(horizontal="center", vertical="center")
    orientacion_alineacion_celda = Alignment(horizontal="center", vertical="center", textRotation=90)


    ws['A1'] = 'Disciplina'
    ws['B1'] = 'Código de proyecto'
    ws['C1'] = 'Fecha de registro'
    ws['D1'] = 'Temporada'
    ws['E1'] = 'Unidad de vegetación'
    ws['F1'] = 'Estación de muestreo'
    ws['G1'] = 'Unidad de muestreo'
    ws['H1'] = 'Subunidad de muestreo'
    ws['I1'] = 'Método de muestreo'
    ws['J1'] = 'Tamaño de la unidad de muestreo'
    ws['K1'] = 'Especie dominante'
    ws['L1'] = 'X (NORTE)'
    ws['M1'] = 'Y (ESTE)'
    ws['N1'] = 'Proyección'
    ws['O1'] = 'Altitud (msnm)'
    ws['P1'] = 'Datum'
    ws['Q1'] = 'Altura total de la muestra (m)'
    ws['R1'] = 'Área de la muestra (l x a)'
    ws['S1'] = 'Profundidad total de calicata (m)'
    ws['T1'] = 'altura de la vegetación '
    ws['U1'] = 'Altura de la materia organica'
    ws['V1'] = 'Estado de conservación del hábitat'
    ws['W1'] = 'Colector de registro'
    ws['X1'] = 'FOTO'
    ws['Y1'] = 'Observación de IMPACTO'

    abc = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y']

    for col in range (1,26):
        cd = ws.cell(row=1,column=col)
        cd.border = borde_celda_1
        cd.fill = fondo_celda_cabeza
        cd.font = fondo_texto_cabeza
        cd.alignment = orientacion_alineacion_celda
        ws.column_dimensions[abc[col-1]].width=11
        ws.row_dimensions[1].height = 90

    for col in range(1, 26):
        cd = ws.cell(2,column=col)
        cd.border = borde_celda_2

    cont=2

    ws.cell(row=cont,column=1).value = dato.disciplina
    ws.cell(row=cont,column=2).value = dato.codigo_proyecto
    ws.cell(row=cont,column=3, value = dato.fecha_registro).number_format = numbers.FORMAT_DATE_XLSX15
    ws.cell(row=cont,column=4).value = dato.temporada
    ws.cell(row=cont,column=5).value = dato.unidad_vegetacion
    ws.cell(row=cont,column=6).value = dato.estacion_muestreo
    ws.cell(row=cont,column=7).value = dato.unidad_muestreo
    ws.cell(row=cont,column=8).value = dato.subunidad_muestreo
    ws.cell(row=cont,column=9).value = dato.metodo_muestreo
    ws.cell(row=cont,column=10).value = dato.tamaño_unidad_muestreo
    ws.cell(row=cont,column=11).value = dato.especie_dominante
    ws.cell(row=cont,column=12).value = dato.x_norte
    ws.cell(row=cont,column=13).value = dato.y_este
    ws.cell(row=cont,column=14).value = dato.proyeccion
    ws.cell(row=cont,column=15).value = dato.altitud_msnm
    ws.cell(row=cont,column=16).value = dato.datum
    ws.cell(row=cont,column=17).value = dato.altura_total_muestra
    ws.cell(row=cont,column=18).value = dato.area_muestra
    ws.cell(row=cont,column=19).value = dato.profundidad_total_calicata
    ws.cell(row=cont,column=20).value = dato.altura_vegetacion
    ws.cell(row=cont,column=21).value = dato.altura_materia_organica
    ws.cell(row=cont,column=22).value = dato.estado_conservacion_habitat
    ws.cell(row=cont,column=23).value = dato.colector_registro
    ws.cell(row=cont,column=24).value = dato.FOTO
    ws.cell(row=cont,column=25).value = dato.observacion_impacto

    nombre_archivo ="Vegetacion_Compensacion_Amb_Unico.xlsx"
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

#-----------------------------------------------------------------------------------------------------------------------
#===================== Formularios_unicos ================================

#form_unico_avez

def formulario_unic_Aves_chinalco(request,id):
    aves_unic=BD_Integrada_Aves_Chinalco_Seca_2018.objects.get(id=id)
    return render(request, 'biologia/formularios_unicos/formulario_unico_Aves_chinalco_BD.html',
                  {'aves_unic':aves_unic})

def formulario_unic_aves_coordenadas_int(request,id):
    aves_coors_unic=CoordenadasIntegradasAves2018.objects.get(id=id)
    return render(request, 'biologia/formularios_unicos/form_unico_Aves_coordenadas_int.html',
                  {'aves_coors_unic':aves_coors_unic})

def formulario_unic_Aves_chinalco_fotos(request,id):
    aves_fotos_unic=Fotos_Integrado_Aves_Chinalco_Seca_2018.objects.get(id=id)
    return render(request, 'biologia/formularios_unicos/form_unico_Aves_Fotos_int.html',
                  {'aves_fotos_unic':aves_fotos_unic})

#form_unico_Reptiles_Amfibios

def formulario_unic_Reptiles_anfibios_BD(request, id):
    reptiles_bd_anfibios = BD_anfibios_reptiles_BD.objects.get(id=id)
    return render(request, 'biologia/formularios_unicos/form_unico_BD_BD_anfibios_reptiles.html',
                    {'reptiles_bd_anfibios':reptiles_bd_anfibios})

def formulario_unic_Reptiles_anfibios_BD_fotos(request, id):
    reptiles_fotos_anfibios = BD_fotos_BD.objects.get(id=id)
    return render(request, 'biologia/formularios_unicos/form_unico_BD_BDFotos.html',
                    {'reptiles_fotos_anfibios':reptiles_fotos_anfibios})

def formulario_unic_Reptiles_anfibios_coordenadas(request, id):
    reptiles_coordenadas_anfibios = Coordenadas_BD.objects.get(id=id)
    return render(request, 'biologia/formularios_unicos/form_unico_BD_Coordenadas.html',
                    {'reptiles_coordenadas_anfibios':reptiles_coordenadas_anfibios})

def formulario_unic_Reptiles_anfibios_BD_hoja1(request,id):
    hoja1_bd_unic = hoja1_BD.objects.get(id=id)
    return render(request, 'biologia/formularios_unicos/form_unico_BD_hoja1.html',
                    {'hoja1_bd_unic':hoja1_bd_unic})

#form_unico_mamiferos

def formulario_unic_Mamiferos_BD(request,id):
    mamiferos_bd_unic = BD_mamiferos_BD_Mamiferos.objects.get(id=id)
    return render(request, 'biologia/formularios_unicos/form_unico_BDMamiferos_BDMamiferos.html',
                    {'mamiferos_bd_unic':mamiferos_bd_unic})

def formulario_unic_Mamiferos_bitacora(request,id):
    bit_mamiferos_unic = bitacora_BD_Mamiferos.objects.get(id=id)
    return render(request, 'biologia/formularios_unicos/form_unico_BDMamiferos_bitacora.html',
                    {'bit_mamiferos_unic':bit_mamiferos_unic})

def formulario_unic_Mamiferos_coordenadas(request,id):
    mamiferos_coordenadas_uni = Coordenadas_reporte_BD_Mamiferos.objects.get(id=id)
    return render(request, 'biologia/formularios_unicos/form_unico_BDMamiferos_coordenadas_rep.html',
                    {'mamiferos_coordenadas_uni':mamiferos_coordenadas_uni})

def formulario_unic_Mamiferos_ezfuerzo_muestreo(request, id):
    es_muestreo_unic = Esfuerzo_Muestreo_BD_Mamiferos.objects.get(id=id)
    return render(request, 'biologia/formularios_unicos/form_unico_BDMamiferos_esfuerzo_muestreo.html',
                    {'es_muestreo_unic':es_muestreo_unic})

#form_unico_vegetales

def formulario_unic_Vegeta_BD(request,id):
    vegetacion_bd_uni = Base_Datos_Vegetacion_2018.objects.get(id=id)
    return render(request, 'biologia/formularios_unicos/form_unico_Vegetacion_BD.html',
                    {'vegetacion_bd_uni':vegetacion_bd_uni})

def formulario_unic_Vegeta_compensacion(request, id):
    vegetacion_compensacion_uni = Compensacion_Ambiental_Vegetacion_2018.objects.get(id=id)
    return render(request, 'biologia/formularios_unicos/form_unico_Vegetacion_compensacion_amb.html',
                        {'vege_compen_uni':vegetacion_compensacion_uni})

def formulario_unic_Vegeta_esfuerzo_muestreo(request, id):
    vegetacion_esfuerzo_uni = Esfuerzo_Muestreo_BD_Vegetacion_2018.objects.get(id=id)
    return render(request, 'biologia/formularios_unicos/form_unico_Vegetacion_esfuerzo_muestreo.html',
                        {'vege_esfuerzo_uni':vegetacion_esfuerzo_uni})

def formulario_unic_Vegeta_hoja3(request, id):
    vegetacion_hoja3_uni = Hoja3_BD_Vegetacion_2018.objects.get(id=id)
    return render(request, 'biologia/formularios_unicos/form_unico_Vegetacion_hoja3.html',
                        {'vege_hoja3_unic':vegetacion_hoja3_uni})

#-----------------------------------------------------------------------------------------------------------------------

#Prueba para coordenadas_BD

